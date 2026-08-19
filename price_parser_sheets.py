import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime
import os
import re
import math
import json
import hashlib
import traceback
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode, urljoin
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
# Подавляем только известный шумный DeprecationWarning от openpyxl/pandas,
# а не все предупреждения подряд — иначе можно пропустить реальную проблему
warnings.filterwarnings('ignore', category=DeprecationWarning)

# Настройка логирования. Лог-файл ротируется, чтобы не расти бесконечно
# при регулярных запусках парсера.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('price_parser.log', maxBytes=5 * 1024 * 1024, backupCount=3, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==========================================================================
# Справочные данные для эвристик парсинга.
# Раньше эти списки были разбросаны по телам методов — вынесены сюда,
# чтобы их можно было править в одном месте, не копаясь в коде парсера.
# ==========================================================================

# Цвета, которые ищем в названии товара (extract_color_from_product_name)
COLOR_KEYWORDS = [
    'прозрачный', 'матовый', 'белый', 'цветной',
    'бронза', 'молочный', 'голубой', 'зеленый',
    'красный', 'желтый', 'черный', 'синий',
    'оранжевый', 'фиолетовый', 'коричневый',
]

# Сокращения/варианты написания цвета -> полное название
COLOR_VARIATIONS = {
    'мат': 'матовый',
    'прозр': 'прозрачный',
    'бел': 'белый',
    'цвет': 'цветной',
}

# Тип материала по ключевому слову в названии (extract_material_type_from_product_name)
MATERIAL_TYPE_KEYWORDS = {
    'поликарбонат': 'поликарбонат',
    'polygal': 'поликарбонат',
    'оргстекло': 'оргстекло',
    'орг.стекло': 'оргстекло',
    'акрил': 'акрил',
    'пластик': 'пластик',
    'пвх': 'пвх',
    'пенопласт': 'пенопласт',
    'дерево': 'дерево',
    'металл': 'металл',
    'алюминий': 'алюминий',
    'сталь': 'сталь',
    'стекло': 'стекло',
}

# Известные бренды/марки (extract_brand_from_product_name)
KNOWN_BRANDS = [
    'polygal', 'plazcryl', 'акрилекс', 'моногал', 'легпром',
    'полигаль', 'плазкрил', 'akrilux', 'monogal',
]

# Тип продукта по ключевому слову (extract_product_type_from_product_name)
PRODUCT_TYPE_KEYWORDS = {
    'сотовый': 'сотовый',
    'монолитный': 'монолитный',
    'листовой': 'листовой',
    'профилированный': 'профилированный',
    'ячеистый': 'ячеистый',
    'сот': 'сотовый',
    'монолит': 'монолитный',
}

# Общие слова без смысловой нагрузки, исключаемые из ключевых слов
# (extract_important_keywords)
KEYWORD_STOP_WORDS = {
    'на', 'и', 'в', 'с', 'для', 'по', 'из', 'от', 'до',
    'мм', 'см', 'м', 'кг', 'г', 'л', 'шт', 'упак',
    'лист', 'листа', 'листов', 'пластина', 'плита',
}

# Диапазоны "разумной" цены по товару (is_reasonable_price).
# Правила проверяются по порядку, побеждает первое, где ВСЕ keywords
# встречаются в названии товара (в нижнем регистре).
PRODUCT_PRICE_RANGES = [
    {
        'keywords': ('поликарбонат', 'polygal', 'сотовый'),
        'min': 5000, 'max': 20000,
    },
    {
        'keywords': ('орг.стекло', 'plazcryl'),
        'min': 5000, 'max': 30000,
    },
    {
        'keywords': ('композитная панель',),
        'min': 3000, 'max': 15000,
    },
]
# Диапазон по умолчанию, если ни одно правило выше не подошло
DEFAULT_PRICE_RANGE = (100, 50000)
# Диапазон для случаев, когда название товара вообще не передано
UNKNOWN_PRODUCT_PRICE_RANGE = (100, 100000)

# Параметры трекинга рекламы/поисковиков, которые иногда попадают в ссылку
# при копировании её из результатов поиска или рекламных кампаний
# (например, https://site.ru/product/?ysclid=... после перехода из Яндекса).
# Они не относятся к самому товару и их стоит убирать перед сохранением
# и запросом ссылки — сайт отдаёт ту же страницу и без них.
TRACKING_QUERY_PARAMS = {
    'ysclid', 'yclid', 'gclid', 'fbclid', 'msclkid', '_openstat',
    'utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content',
}

# Фразы в <title>/<h1>, по которым распознаём "мягкий" 404 — страницу вида
# "товар не найден", отданную с HTTP-статусом 200 (см. looks_like_error_page)
ERROR_PAGE_MARKERS = (
    'страница не найдена', 'страницы не существует', 'страница не существует',
    'товар не найден', 'товар не существует', 'page not found', 'not found',
    'ошибка 404', '404 not found',
)


def strip_tracking_params(url):
    """
    Убирает известные рекламные/поисковые трекинг-параметры (ysclid, utm_*
    и т.п.) из URL, сохраняя остальные query-параметры как есть — некоторые
    сайты используют их для выбора конкретного товара/варианта, и их
    удалять нельзя.
    """
    if not url:
        return url
    try:
        parts = urlsplit(url)
        if not parts.query:
            return url
        filtered = [
            (key, value) for key, value in parse_qsl(parts.query, keep_blank_values=True)
            if key not in TRACKING_QUERY_PARAMS
        ]
        new_query = urlencode(filtered)
        return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))
    except Exception:
        return url


def safe_str(value, default=''):
    """
    Безопасное преобразование значения в строку.
    Обрабатывает None, NaN, float и другие типы.
    """
    if pd.isna(value):  # Проверка на NaN (работает для float('nan') и pd.NA)
        return default
    if value is None:
        return default
    try:
        # Проверка на NaN: NaN не равен самому себе
        if isinstance(value, float) and value != value:
            return default
        return str(value)
    except Exception:
        return default


class PriceParserWithSheets:
    def __init__(self, excel_file_path, sheet_name="Прайс-лист"):
        """
        Инициализация парсера для работы с конкретным листом Excel
        """
        self.excel_file = excel_file_path
        self.sheet_name = sheet_name
        self.df = None
        self.driver = None
        self.results = []
        self.rounding_mode = 'ceil'  # Режим округления по умолчанию
        self.user_selections = {}  # Словарь для хранения выборов пользователя
        # Финальный URL и статус последнего запроса через get_with_requests
        # (см. этот метод — используется для самолечения "протухших" ссылок)
        self.last_fetched_url = None
        self.last_status_code = None
        
        # Загружаем сохраненные выборы пользователя при инициализации
        self.load_user_selections()
        
        # Улучшенная конфигурация для разных сайтов
        self.site_configs = {
            'expo-torg.ru': {
                'name': 'Экспо-Торг',
                'method': 'requests',
                'headers': {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                }
            },
            'bestly.ru': {
                'name': 'Bestly',
                'method': 'selenium',  # Используем Selenium для Bestly
                'selenium_wait': 20,  # Увеличенное время ожидания для Bestly
                'dynamic_content': True,  # Указываем, что контент динамический
                'use_selenium_context_search': True,  # Использовать контекстный поиск через Selenium
                'price_selectors': [
                    '.item_price',  # Основной селектор для bestly.ru из данных
                    '.price', '.product-price', '.woocommerce-Price-amount',
                    '.product-item-price', '.regular-price', '.current-price',
                    '.product-card__price', '.catalog-item-price',
                    'span.price', 'div.price', 'p.price',
                    '[data-price]', '[itemprop="price"]',
                    '.product-card-price', '.item-price', '.price-value',
                    '.product_price', '.card-price', '.catalog-price',
                    '.js-product-price', '.product__price', '.price__current',
                    '.product-price-current', '.product-price__value',
                    '.price-value', '.price-new', '.price-old',
                    '.b-product-price__current',  # Новые селекторы для Bestly
                    '.product-card__price-current',
                    '.price_item', '.price-block',
                    'div[class*="price"]', 'span[class*="price"]',
                    '.price-replace-bottom'  # Добавляем специфичный селектор для Bestly
                ],
                'headers': {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                },
                'card_selectors': [  # Селекторы для карточек товаров
                    '.catalog-item', '.product-item', '.item', 
                    'div[class*="item"]', 'div[class*="product"]',
                    '.catalog-section', '.product-card', '.goods-item'
                ],
                'special_parsers': {  # Специальные парсеры для определенных страниц
                    'org_steklo_plazcryl': 'parse_orgsteklo_table'
                }
            },
            'donalum.ru': {
                'name': 'Доналюм',
                'method': 'requests',
                'headers': {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                }
            },
            'expocomplete.ru': {
                'name': 'ExpoComplete',
                'method': 'requests',
                'headers': {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                }
            }
        }
    
    def load_user_selections(self):
        """Загрузка сохраненных выборов пользователя из файла"""
        try:
            selections_file = 'user_selections.json'
            if os.path.exists(selections_file):
                with open(selections_file, 'r', encoding='utf-8') as f:
                    self.user_selections = json.load(f)
                logger.info(f"Загружено {len(self.user_selections)} сохраненных выборов пользователя")
        except Exception as e:
            logger.warning(f"Не удалось загрузить сохраненные выборы: {e}")
            self.user_selections = {}
    
    def save_user_selections(self):
        """Сохранение выборов пользователя в файл"""
        try:
            selections_file = 'user_selections.json'
            with open(selections_file, 'w', encoding='utf-8') as f:
                json.dump(self.user_selections, f, ensure_ascii=False, indent=2)
            logger.info(f"Сохранено {len(self.user_selections)} выборов пользователя")
        except Exception as e:
            logger.error(f"Ошибка сохранения выборов пользователя: {e}")
    
    def save_user_selection(self, product_url, selector, text, price):
        """Сохранение выбора пользователя для конкретного товара"""
        try:
            # Создаем стабильный ключ на основе URL и текста элемента.
            # Важно использовать hashlib, а не встроенный hash(): для строк
            # он рандомизируется на каждый запуск процесса (PYTHONHASHSEED),
            # из-за чего один и тот же выбор сохранялся бы под новым ключом
            # при каждом запуске скрипта, и user_selections.json бесконтрольно рос.
            digest = hashlib.md5((selector + text).encode('utf-8')).hexdigest()
            key = f"{product_url}_{digest}"
            
            self.user_selections[key] = {
                'url': product_url,
                'selector': selector,
                'text': text[:200],  # Сохраняем только первые 200 символов
                'price': price,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Сохраняем в файл
            self.save_user_selections()
            logger.info(f"Сохранен выбор пользователя для {product_url}")
            return True
        except Exception as e:
            logger.error(f"Ошибка сохранения выбора пользователя: {e}")
            return False
    
    def get_user_selection(self, product_url):
        """Получение сохраненного выбора пользователя для товара"""
        try:
            # Ищем все сохраненные выборы для этого URL
            selections = []
            for key, value in self.user_selections.items():
                if value['url'] == product_url:
                    selections.append(value)
            
            # Возвращаем последний выбор (самый актуальный)
            if selections:
                # Сортируем по timestamp (новые сначала)
                selections.sort(key=lambda x: x['timestamp'], reverse=True)
                return selections[0]
            
            return None
        except Exception as e:
            logger.error(f"Ошибка получения сохраненного выбора: {e}")
            return None
    
    def set_rounding_mode(self, mode='ceil'):
        """Установка режима округления цен"""
        valid_modes = ['ceil', 'floor', 'nearest', 'nearest_decimal', 'no_decimal']
        if mode in valid_modes:
            self.rounding_mode = mode
            logger.info(f"Установлен режим округления: {mode}")
            return True
        else:
            logger.error(f"Неверный режим округления. Допустимые значения: {valid_modes}")
            return False
    
    def round_price(self, price, mode=None):
        """Округление цены в зависимости от режима"""
        if mode is None:
            mode = self.rounding_mode
        
        try:
            if mode == 'ceil':
                # Округление в большую сторону до 2 знаков после запятой
                return math.ceil(price * 100) / 100
            elif mode == 'floor':
                # Округление в меньшую сторону
                return math.floor(price * 100) / 100
            elif mode == 'nearest':
                # Округление до ближайшего целого числа
                return round(price)
            elif mode == 'nearest_decimal':
                # Округление до 2 знаков после запятой
                return round(price, 2)
            elif mode == 'no_decimal':
                # Округление до целого числа в большую сторону
                return math.ceil(price)
            else:
                # По умолчанию округляем в большую сторону до 2 знаков
                return math.ceil(price * 100) / 100
        except Exception as e:
            logger.error(f"Ошибка округления цены: {e}")
            return round(price, 2)  # Возвращаем просто округленную цену при ошибке

    # ----------------------------------------------------------------
    # Общие вспомогательные методы для поиска толщины в тексте.
    # Раньше практически идентичный набор regex-паттернов и цикл
    # числового сравнения были продублированы более чем в 7 местах
    # (find_best_match_by_thickness, calculate_match_score,
    # debug_row_info, check_thickness_in_text,
    # universal_table_parsing_bestly, parse_orgsteklo_table_improved,
    # parse_bestly_orgsteklo_table). Теперь это одно место.
    # ----------------------------------------------------------------

    def _numbers_in_text(self, text):
        """Извлекает все числа из текста в виде float (запятая -> точка)."""
        numbers = []
        if not text:
            return numbers
        for num_str in re.findall(r'\d+(?:[.,]\d+)?', text):
            try:
                numbers.append(float(num_str.replace(',', '.')))
            except ValueError:
                continue
        return numbers

    def _thickness_regex_patterns(self, thickness, mode='boundary'):
        """
        Возвращает regex-паттерны для поиска толщины в тексте.
          'exact'    — ячейка таблицы целиком (^...$) плюс варианты с
                       границей слова; используется при разборе табличных
                       ячеек, где ожидается только значение толщины.
          'boundary' — варианты с границей слова (\\b...\\b), включая
                       "мм"/"mm", чтобы толщина "4" не совпадала как часть
                       более длинного числа (например "14мм").
        """
        escaped = re.escape(str(thickness))
        if mode == 'exact':
            return [
                f'^{escaped}\\s*мм$',
                f'^{escaped}\\s*mm$',
                f'^{escaped}$',
                f'\\b{escaped}\\s*мм\\b',
                f'\\b{escaped}\\s*mm\\b',
                f'\\b{escaped}\\b',
            ]
        # 'boundary' (по умолчанию)
        return [
            f'\\b{escaped}\\s*мм\\b',
            f'\\b{escaped}\\b',
            f'\\b{escaped}\\s*mm\\b',
        ]

    def _thickness_regex_match(self, text, thickness, mode='boundary'):
        """Проверяет текст только по regex-паттернам толщины (без числового fallback)."""
        if not text or not thickness:
            return False
        text_lower = text.lower()
        for pattern in self._thickness_regex_patterns(thickness, mode=mode):
            if re.search(pattern, text_lower, re.IGNORECASE):
                return True
        return False

    def _text_matches_thickness(self, text, thickness, mode='boundary', tolerance=0.1):
        """
        Проверяет, соответствует ли текст указанной толщине: сначала по
        regex-паттернам (см. _thickness_regex_patterns), затем, если не
        нашли, по числовому совпадению с заданным допуском.
        """
        if not text or not thickness:
            return False

        if self._thickness_regex_match(text, thickness, mode=mode):
            return True

        try:
            thickness_val = float(thickness)
        except (TypeError, ValueError):
            return False

        for num in self._numbers_in_text(text.lower()):
            if abs(num - thickness_val) < tolerance:
                return True

        return False

    def extract_color_from_product_name(self, product_name):
        """
        Извлекает цвет из названия товара.
        Примеры:
        - "Орг.стекло PLAZCRYL прозрачный 4мм" → "прозрачный"
        - "Орг.стекло 4мм прозрачный" → "прозрачный"
        - "Орг.стекло матовое 4мм" → "матовый"
        - "Орг.стекло белый 4мм" → "белый"
        """
        try:
            if not product_name:
                return None
            
            # Убираем лишние пробелы и приводим к нижнему регистру
            clean_name = product_name.lower().strip()

            # Ищем цвет в названии
            for color in COLOR_KEYWORDS:
                if color in clean_name:
                    logger.info(f"Извлечен цвет из названия '{product_name}': {color}")
                    return color

            # Также проверяем возможные вариации
            for variation, full_color in COLOR_VARIATIONS.items():
                if variation in clean_name:
                    logger.info(f"Извлечен цвет из вариации '{product_name}': {full_color}")
                    return full_color

            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения цвета: {e}")
            return None
    
    def init_selenium_driver(self):
        """Инициализация Selenium драйвера с улучшенными настройками для Bestly"""
        try:
            logger.info("Попытка инициализации Selenium драйвера...")
            
            chrome_options = Options()
            # Раскомментируйте следующую строку, если хотите видеть браузер
            # chrome_options.add_argument('--headless')  # Закомментируйте для отладки
            
            # Базовые настройки
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            
            # Улучшенные настройки для сайтов с динамическим контентом
            chrome_options.add_argument('--disable-web-security')
            chrome_options.add_argument('--allow-running-insecure-content')
            chrome_options.add_argument('--disable-features=VizDisplayCompositor')
            
            # Отключаем автоматизацию
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_experimental_option("prefs", {
                "profile.default_content_setting_values.notifications": 2,
                "profile.managed_default_content_settings.images": 1,  # Разрешаем картинки
                "profile.default_content_setting_values.popups": 2,
                "profile.default_content_setting_values.javascript": 1,  # Включаем JavaScript
                "profile.default_content_setting_values.cookies": 1,  # Разрешаем cookies
            })
            
            # Добавляем User-Agent
            chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            
            try:
                # Пробуем использовать webdriver_manager
                logger.info("Пробуем использовать webdriver_manager...")
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
            except Exception as e:
                logger.warning(f"webdriver_manager не сработал: {e}")
                logger.info("Пробуем использовать драйвер из PATH...")
                # Пробуем использовать драйвер из PATH
                self.driver = webdriver.Chrome(options=chrome_options)
            
            # Устанавливаем таймауты
            self.driver.set_page_load_timeout(120)  # Увеличиваем таймаут для Bestly
            self.driver.implicitly_wait(20)  # Увеличиваем неявное ожидание
            
            # Устанавливаем размер окна
            self.driver.set_window_size(1920, 1080)
            
            # Выполняем скрипт для скрытия автоматизации
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.driver.execute_script("Object.defineProperty(navigator, 'languages', {get: () => ['ru-RU', 'ru', 'en-US', 'en']})")
            self.driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]})")
            
            logger.info("✓ Selenium драйвер успешно инициализирован")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при инициализации Selenium: {e}")
            return False
    
    def close_selenium_driver(self):
        """Закрытие Selenium драйвера"""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("Selenium драйвер закрыт")
            except Exception:
                pass
            finally:
                self.driver = None
    
    def load_excel_data(self):
        """Загрузка данных из Excel файла с указанного листа"""
        try:
            if not os.path.exists(self.excel_file):
                logger.error(f"Файл {self.excel_file} не найден!")
                return False
            
            # Загружаем Excel файл
            xl = pd.ExcelFile(self.excel_file, engine='openpyxl')
            
            # Проверяем наличие нужного листа
            if self.sheet_name not in xl.sheet_names:
                logger.error(f"Лист '{self.sheet_name}' не найден в файле!")
                logger.info(f"Доступные листы: {xl.sheet_names}")
                return False
            
            # Читаем данные с нужного листа
            self.df = xl.parse(self.sheet_name)
            
            # Переименовываем колонки для удобства
            column_mapping = {}
            for col in self.df.columns:
                col_str = str(col).strip()
                # Приводим названия колонок к стандартному виду
                if 'название' in col_str.lower():
                    column_mapping[col] = 'Название'
                elif 'единица' in col_str.lower():
                    column_mapping[col] = 'Единица'
                elif 'цена' in col_str.lower() and 'за м2' not in col_str.lower():
                    column_mapping[col] = 'Цена'
                elif 'url' in col_str.lower() or 'ссылка' in col_str.lower():
                    column_mapping[col] = 'URL'
                elif 'селектор' in col_str.lower():
                    column_mapping[col] = 'Селектор'
                elif 'характеристика' in col_str.lower():
                    column_mapping[col] = 'Характеристика'
                elif 'дата' in col_str.lower() and 'обнов' in col_str.lower():
                    column_mapping[col] = 'Дата обновления'
            
            # Переименовываем колонки
            if column_mapping:
                self.df.rename(columns=column_mapping, inplace=True)
            
            # Проверяем наличие необходимых колонок
            required_columns = ['Название', 'URL']
            missing_columns = [col for col in required_columns if col not in self.df.columns]
            
            if missing_columns:
                logger.error(f"Отсутствуют необходимые колонки: {missing_columns}")
                logger.info(f"Найденные колонки: {list(self.df.columns)}")
                return False
            
            # Добавляем колонки, если их нет
            if 'Селектор' not in self.df.columns:
                self.df['Селектор'] = ''
            
            if 'Характеристика' not in self.df.columns:
                self.df['Характеристика'] = ''
            
            if 'Дата обновления' not in self.df.columns:
                self.df['Дата обновления'] = ''
            
            if 'Цена' not in self.df.columns:
                self.df['Цена'] = ''
            
            logger.info(f"Загружен файл: {self.excel_file}")
            logger.info(f"Лист: {self.sheet_name}")
            logger.info(f"Колонки: {list(self.df.columns)}")
            logger.info(f"Количество товаров: {len(self.df)}")
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка загрузки файла Excel: {e}")
            traceback.print_exc()
            return False
    
    def extract_domain(self, url):
        """Извлечение домена из URL"""
        try:
            if pd.isna(url) or not isinstance(url, str):
                return ''
            
            url = str(url).strip()
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            
            # Извлекаем домен
            pattern = r'https?://(?:www\.)?([^/]+)'
            match = re.search(pattern, url)
            if match:
                domain = match.group(1).lower()
                # Удаляем www.
                if domain.startswith('www.'):
                    domain = domain[4:]
                return domain
            return ''
        except Exception:
            return ''

    def looks_like_error_page(self, html):
        """
        Определяет "мягкий" 404 — когда сайт вместо HTTP-статуса 404
        отдаёт страницу-заглушку "товар не найден" с кодом 200 (типичная
        конфигурация SPA/CMS). get_with_requests() в этом случае не может
        распознать проблему по статус-коду, а без этой проверки текст
        заглушки уходил в поиск цены (например, число "404" из заголовка
        "Страница не найдена (404 Not Found)" принималось за цену).

        Проверяет только <title> и первый <h1> — не весь текст страницы,
        чтобы не срабатывать на обычные товары, у которых "404" может
        случайно встретиться в артикуле где-то в теле страницы.
        """
        if not html:
            return False
        try:
            soup = BeautifulSoup(html, 'html.parser')
            texts = []
            title = soup.find('title')
            if title:
                texts.append(title.get_text(strip=True).lower())
            h1 = soup.find('h1')
            if h1:
                texts.append(h1.get_text(strip=True).lower())
            combined = ' '.join(texts)
            return any(marker in combined for marker in ERROR_PAGE_MARKERS)
        except Exception:
            return False

    # ----------------------------------------------------------------
    # Восстановление "умершей" ссылки на товар через раздел каталога.
    # Если ссылка на конкретный товар больше не работает (404 или
    # страница-заглушка), сама категория обычно продолжает существовать
    # — пробуем открыть её и найти в списке товар с максимально похожим
    # названием, вместо того чтобы просто сообщать "ссылка не работает".
    # ----------------------------------------------------------------

    def derive_category_url(self, product_url):
        """
        Отрезает от ссылки на товар последний сегмент пути, получая
        ссылку на его категорию каталога.
        Например: .../catalog/plastik/bumazhno-sloistyy-plastik-hpl/bsp-.../
                -> .../catalog/plastik/bumazhno-sloistyy-plastik-hpl/
        """
        try:
            parts = urlsplit(product_url)
            segments = [s for s in parts.path.split('/') if s]
            if len(segments) < 2:
                return None
            category_path = '/' + '/'.join(segments[:-1]) + '/'
            return urlunsplit((parts.scheme, parts.netloc, category_path, '', ''))
        except Exception:
            return None

    def get_category_listing_html(self, category_url, max_clicks=25):
        """
        Загружает страницу категории каталога и кликает "Показать ещё"
        (подгрузка товаров через JS без смены ссылки), пока список не
        перестанет расти или кнопка не пропадёт — чтобы получить полный
        список товаров категории, а не только первую партию.
        """
        if not self.driver:
            if not self.init_selenium_driver():
                logger.error("Не удалось инициализировать Selenium для загрузки категории")
                return None
        try:
            logger.info(f"Открываем категорию: {category_url}")
            self.driver.get(category_url)
            time.sleep(3)

            previous_count = -1
            for _ in range(max_clicks):
                current_count = len(self.driver.find_elements(By.CSS_SELECTOR, '.product-item'))
                if current_count == previous_count:
                    break
                previous_count = current_count
                try:
                    buttons = self.driver.find_elements(
                        By.XPATH,
                        "//*[contains(text(), 'Показать ещё') or contains(text(), 'Показать еще') "
                        "or contains(text(), 'Загрузить ещё') or contains(text(), 'Загрузить еще')]"
                    )
                    if not buttons:
                        break
                    buttons[0].click()
                    time.sleep(2)
                except Exception as e:
                    logger.debug(f"Не удалось кликнуть 'Показать ещё': {e}")
                    break

            return self.driver.page_source
        except Exception as e:
            logger.error(f"Ошибка загрузки категории {category_url}: {e}")
            return None

    def parse_category_product_cards(self, html, base_url):
        """
        Извлекает из HTML страницы категории список товаров (ссылка +
        название). Рассчитано на разметку вида expo-torg.ru:
            <div class="product-item">
                <div class="product-item-title"><a href="..." title="...">Название</a></div>
        """
        cards = []
        try:
            soup = BeautifulSoup(html, 'html.parser')
            for item in soup.select('.product-item'):
                link = item.select_one('.product-item-title a') or item.select_one('a[href]')
                if not link or not link.get('href'):
                    continue
                name = link.get('title') or link.get_text(strip=True)
                if not name:
                    continue
                cards.append({'url': urljoin(base_url, link['href']), 'name': name})
        except Exception as e:
            logger.error(f"Ошибка разбора карточек категории: {e}")
        return cards

    def _match_tokens(self, text):
        """
        Токенизирует название товара для сопоставления кандидатов в
        find_best_category_match. В отличие от extract_important_keywords,
        НЕ отбрасывает числовые токены: коды декора и размеры (например,
        "3155", "4200", "1860") часто оказываются единственным отличием
        между очень похожими товарами одной категории — например,
        "Компакт-плита 3150 ... Светло-серый" и "3155 ... Антрацит серый"
        отличаются именно кодом и цветом, а не структурой названия.
        """
        if not text:
            return set()
        words = re.findall(r'[а-яa-z0-9]{2,}', text.lower())
        return {w for w in words if w not in KEYWORD_STOP_WORDS}

    def find_best_category_match(self, candidates, target_name, min_ratio=0.7, min_margin=0.1):
        """
        Ищет среди candidates ({'url','name'}) наиболее вероятное
        соответствие target_name. Числовые токены (коды/размеры) из
        целевого названия ОБЯЗАНЫ присутствовать у кандидата целиком —
        это самый надёжный признак различия между похожими товарами;
        остальные слова (материал, цвет, бренд) используются для ранжирования.

        Возвращает (кандидат, список_с_баллами) при уверенном совпадении,
        иначе (None, список_с_баллами) — лучше не найти замену вовсе, чем
        подставить случайно похожий, но другой товар.
        """
        target_tokens = self._match_tokens(target_name)
        if not target_tokens or not candidates:
            return None, []

        target_numbers = {t for t in target_tokens if t.isdigit()}

        scored = []
        for candidate in candidates:
            candidate_tokens = self._match_tokens(candidate['name'])
            if not candidate_tokens:
                continue
            if target_numbers and not target_numbers.issubset(candidate_tokens):
                continue
            overlap = target_tokens & candidate_tokens
            union = target_tokens | candidate_tokens
            ratio = len(overlap) / len(union) if union else 0
            scored.append({'candidate': candidate, 'ratio': ratio})

        scored.sort(key=lambda x: x['ratio'], reverse=True)
        if not scored:
            return None, scored

        best = scored[0]
        second_ratio = scored[1]['ratio'] if len(scored) > 1 else 0

        if best['ratio'] >= min_ratio and (len(scored) == 1 or (best['ratio'] - second_ratio) >= min_margin):
            return best['candidate'], scored

        return None, scored

    def recover_dead_link(self, dead_url, product_name):
        """
        Пытается найти новый адрес товара, если старая ссылка "умерла".
        Возвращает URL совпадения или None, если уверенного совпадения
        не нашлось.
        """
        domain = self.extract_domain(dead_url)
        if 'bestly.ru' in domain:
            # У bestly.ru товары лежат плоско по одному .html на слаг, без
            # вложенных категорий вида expo-torg.ru — обрезка до "каталога"
            # дала бы весь сайт целиком, а не узкий список для сравнения.
            return None

        category_url = self.derive_category_url(dead_url)
        if not category_url:
            logger.warning(f"Не удалось определить категорию для {dead_url}")
            return None

        logger.info(f"Восстановление ссылки: ищем '{product_name}' в категории {category_url}")
        html = self.get_category_listing_html(category_url)
        if not html:
            logger.warning(f"Не удалось загрузить категорию {category_url}")
            return None

        candidates = self.parse_category_product_cards(html, category_url)
        if not candidates:
            logger.warning(f"В категории {category_url} не найдено ни одной карточки товара")
            return None

        logger.info(f"В категории найдено {len(candidates)} товаров, ищем совпадение...")
        best, scored = self.find_best_category_match(candidates, product_name)

        if best:
            logger.info(f"✓ Найдена замена ссылки: '{best['name']}' -> {best['url']}")
            return best['url']

        for entry in sorted(scored, key=lambda x: x['ratio'], reverse=True)[:3]:
            logger.info(f"  Похожий, но неуверенный вариант ({entry['ratio']:.2f}): {entry['candidate']['name']}")
        logger.warning(f"Не найдено уверенного совпадения для '{product_name}' в категории")
        return None

    def try_recover_and_refetch(self, dead_url, product_name, domain):
        """
        Ищет замену умершей ссылке (recover_dead_link) и сразу же
        загружает найденную страницу через requests. Возвращает
        (новый_url, html) при успехе, иначе (None, None).
        """
        replacement_url = self.recover_dead_link(dead_url, product_name)
        if not replacement_url:
            return None, None

        headers = self.site_configs.get(domain, {}).get('headers', {})
        new_html = self.get_with_requests(replacement_url, headers)
        if not new_html or self.looks_like_error_page(new_html):
            logger.warning(f"Найденная замена ссылки тоже не открылась: {replacement_url}")
            return None, None

        return replacement_url, new_html

    def extract_thickness_from_product_name(self, product_name):
        """
        Извлекает толщину из названия товара.
        Примеры:
        - "Орг.стекло PLAZCRYL 6мм" → "6"
        - "прозрачный102050x3050" → "10"
        - "Орг.стекло 8мм прозрачный" → "8"
        - "прозрачный 6 2050х3050" → "6"
        """
        try:
            if not product_name:
                return None
            
            # Убираем лишние пробелы и приводим к нижнему регистру
            clean_name = product_name.lower().strip()
            
            # Паттерны для поиска толщины (в порядке приоритета)
            patterns = [
                # Явное указание толщины с "мм"
                r'(\d+(?:[.,]\d+)?)\s*мм\b',
                r'(\d+(?:[.,]\d+)?)\s*mm\b',
                
                # Толщина после типа стекла
                r'(?:прозрачный|матовый|белый|цветной)\s*(\d+(?:[.,]\d+)?)',
                
                # Толщина перед размером
                r'(\d+(?:[.,]\d+)?)\s+(?=\d{4}[хx×]\d{4})',
                
                # Любое число из 1-2 цифр, не являющееся частью размера
                r'\b(\d{1,2}(?:[.,]\d+)?)\b(?!\s*\d{4}[хx×])',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, clean_name, re.IGNORECASE)
                if match:
                    thickness_str = match.group(1)
                    # Заменяем запятую на точку
                    thickness_str = thickness_str.replace(',', '.')
                    
                    try:
                        thickness = float(thickness_str)
                        # Проверяем, что толщина в разумных пределах для оргстекла
                        if 0.5 <= thickness <= 50:
                            logger.info(f"Извлечена толщина из названия '{product_name}': {thickness}мм")
                            return str(thickness)
                    except ValueError:
                        continue
            
            # Если не нашли по паттернам, ищем все числа
            all_numbers = re.findall(r'\d+(?:[.,]\d+)?', clean_name)
            logger.info(f"Все найденные числа в названии: {all_numbers}")
            
            for num_str in all_numbers:
                try:
                    num = float(num_str.replace(',', '.'))
                    # Пропускаем размеры (4-5 цифр)
                    if 1000 <= num <= 99999:
                        continue
                    # Проверяем диапазон толщины
                    if 0.5 <= num <= 50:
                        logger.info(f"Извлечена толщина из цифр в названии '{product_name}': {num}мм")
                        return str(num)
                except Exception:
                    continue
            
            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения толщины: {e}")
            return None
            
    def extract_price_from_text(self, text):
        """Извлечение цены из текста с улучшенной логикой и округлением"""
        if not text or pd.isna(text) or not isinstance(text, str):
            return None
        
        try:
            # Очищаем текст
            text = text.strip()
            
            # Убираем неразрывные пробелы и специальные символы
            text = text.replace('\xa0', ' ').replace('\u2009', ' ').replace('\u202f', ' ')
            text = text.replace('€', '').replace('$', '').replace('₽', '').replace('р.', '').replace('руб.', '').replace('RUB', '').replace('RUR', '')
            text = text.replace('₽/л.', '').replace('₽/л', '').replace('р/л.', '').replace('р/л', '').replace('/л.', '').replace('/л', '')
            
            # Удаляем любые упоминания толщины, чтобы цена "6 388.80₽" не воспринималась как толщина 6
            # Ищем и удаляем шаблоны типа "X мм", "Xmm" в начале текста
            text = re.sub(r'^\d+(?:[.,]\d+)?\s*(?:мм|mm)\b', '', text, flags=re.IGNORECASE)
            text = re.sub(r'\b\d+(?:[.,]\d+)?\s*(?:мм|mm)\b', '', text, flags=re.IGNORECASE)
            
            # Паттерны для поиска цен
            patterns = [
                # Формат: 1 234,56 или 1,234.56
                r'([\d\s]+[.,]\d{1,2})',
                # Формат: 1234
                r'([\d\s]+)',
                # Любое число с разделителями тысяч
                r'(\d{1,3}(?:\s\d{3})*(?:[.,]\d{1,2})?)',
                # Просто число
                r'(\d+(?:[.,]\d{1,2})?)',
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    if isinstance(match, tuple):
                        match = match[0]
                    
                    if match:
                        # Очищаем от пробелов и заменяем запятые на точки
                        cleaned = str(match).replace(' ', '').replace(',', '.')
                        
                        # Удаляем лишние точки
                        parts = cleaned.split('.')
                        if len(parts) > 2:
                            cleaned = parts[0] + '.' + ''.join(parts[1:])
                        
                        try:
                            price = float(cleaned)
                            # Проверяем на разумные пределы (от 0.01 до 1 млн)
                            if 0.01 <= price <= 1000000:
                                # Округляем цену согласно установленному режиму
                                rounded_price = self.round_price(price)
                                return rounded_price
                        except ValueError:
                            continue
            
            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения цены: {e}")
            return None
    
    def extract_product_name_from_page(self, html, selector=None):
        """Извлечение названия товара со страницы"""
        try:
            soup = BeautifulSoup(html, 'html.parser')
            
            # Сначала пробуем указанный селектор
            if selector and selector.strip():
                elements = soup.select(selector.strip())
                for element in elements:
                    text = element.get_text(strip=True)
                    if text and len(text) > 3:  # Минимальная длина названия
                        return text
            
            # Ищем по популярным селекторам названий товаров
            name_selectors = [
                'h1', '.product-title', '.product-name', '.item-title', 
                '.card-title', '.product-card__title', '.catalog-item__title',
                '.product__title', '.title', '.product-name', '.product_title',
                '.product-heading', '.product-header', '.product-info__title',
                '.product-card-title', '.item-name', '.product-item-title',
                'span[itemprop="name"]', 'h1[itemprop="name"]',
                '.product-info-title', '.product-details-title'
            ]
            
            for sel in name_selectors:
                elements = soup.select(sel)
                for element in elements:
                    text = element.get_text(strip=True)
                    if text and len(text) > 3:
                        # Проверяем, что это не цена (не содержит только цифры)
                        if not re.match(r'^[\d\s.,]+$', text):
                            return text
            
            # Ищем в заголовках и мета-тегах
            title = soup.find('title')
            if title:
                title_text = title.get_text(strip=True)
                if title_text and len(title_text) > 3:
                    return title_text
            
            # Ищем в мета-тегах
            meta_name = soup.find('meta', {'name': 'description'})
            if meta_name and meta_name.get('content'):
                meta_text = meta_name.get('content')
                if meta_text and len(meta_text) > 10:
                    # Берем первые 100 символов
                    return meta_text[:100]
            
            return None
            
        except Exception as e:
            logger.error(f"Ошибка извлечения названия товара: {e}")
            return None
    
    def find_similar_elements_by_selector(self, html, selector_patterns):
        """
        Находит все элементы, которые соответствуют списку селекторов.
        Возвращает список элементов и их текстовое содержимое.
        """
        try:
            soup = BeautifulSoup(html, 'html.parser')
            all_elements = []
            
            for pattern in selector_patterns:
                try:
                    elements = soup.select(pattern)
                    for element in elements:
                        text = element.get_text(strip=True)
                        if text and len(text) > 1:  # Минимальная длина текста
                            all_elements.append({
                                'element': element,
                                'text': text,
                                'selector': pattern,
                                'html': str(element)[:200]  # Берем первые 200 символов HTML
                            })
                except Exception as e:
                    logger.warning(f"Ошибка при поиске по селектору {pattern}: {e}")
                    continue
            
            # Удаляем дубликаты (по тексту)
            unique_elements = []
            seen_texts = set()
            for elem in all_elements:
                if elem['text'] not in seen_texts:
                    seen_texts.add(elem['text'])
                    unique_elements.append(elem)
            
            logger.info(f"Найдено {len(unique_elements)} уникальных элементов по {len(selector_patterns)} селекторам")
            return unique_elements
        except Exception as e:
            logger.error(f"Ошибка поиска похожих элементов: {e}")
            return []
    
    def find_best_match_by_thickness(self, elements, thickness):
        """
        Ищет лучший элемент по толщине.
        Возвращает элемент, который содержит указанную толщину.
        """
        if not thickness:
            return None
        
        try:
            logger.info(f"Ищем элемент с толщиной {thickness}мм среди {len(elements)} элементов")
            
            candidates = []

            for elem in elements:
                text = elem['text'].lower()

                found = self._text_matches_thickness(text, thickness, tolerance=0.1)

                if found:
                    # Извлекаем цену из элемента
                    price = self.extract_price_from_text(text)
                    if price:
                        elem['price'] = price
                        elem['thickness_found'] = True
                        candidates.append(elem)
                        logger.info(f"Найден кандидат с толщиной {thickness}мм: {text[:100]}... (цена: {price})")
            
            if candidates:
                # Сортируем кандидатов по близости текста к толщине
                def candidate_score(candidate):
                    score = 0
                    text = candidate['text']
                    
                    # Выше балл, если толщина указана с "мм"
                    if f'{thickness}мм' in text or f'{thickness} mm' in text:
                        score += 10
                    
                    # Выше балл, если есть слово "толщина" рядом
                    if 'толщин' in text.lower():
                        score += 5
                    
                    # Выше балл, если цена находится рядом с толщиной в тексте
                    # Ищем позицию толщины и цены в тексте
                    thickness_pos = text.find(thickness)
                    price_pos = text.find(str(candidate.get('price', '')))
                    
                    if thickness_pos != -1 and price_pos != -1:
                        distance = abs(thickness_pos - price_pos)
                        if distance < 50:  # Если цена и толщина близко
                            score += 20 - (distance / 10)
                    
                    return score
                
                candidates.sort(key=candidate_score, reverse=True)
                best_candidate = candidates[0]
                logger.info(f"Выбран лучший кандидат: {best_candidate['text'][:100]}... (цена: {best_candidate['price']})")
                return best_candidate
            
            logger.warning(f"Не найден элемент с толщиной {thickness}мм")
            return None
            
        except Exception as e:
            logger.error(f"Ошибка поиска лучшего элемента по толщине: {e}")
            return None
    
    def find_all_potential_price_elements(self, html):
        """
        Находит все потенциальные элементы с ценами на странице.
        Использует различные стратегии поиска.
        """
        try:
            soup = BeautifulSoup(html, 'html.parser')
            all_elements = []
            
            # 1. Ищем элементы с классами, содержащими price, cost, value и т.д.
            price_patterns = ['price', 'cost', 'value', 'amount', 'стоимость', 'цена', 'руб', '₽']
            for pattern in price_patterns:
                elements = soup.find_all(class_=re.compile(pattern, re.IGNORECASE))
                for elem in elements:
                    text = elem.get_text(strip=True)
                    if text and len(text) > 1:
                        all_elements.append({
                            'element': elem,
                            'text': text,
                            'selector': f'[class*="{pattern}"]',
                            'method': 'class_pattern'
                        })
            
            # 2. Ищем элементы с атрибутами данных, содержащими цены
            data_attrs = ['data-price', 'data-cost', 'data-value']
            for attr in data_attrs:
                elements = soup.find_all(attrs={attr: True})
                for elem in elements:
                    data_value = elem.get(attr)
                    if data_value:
                        all_elements.append({
                            'element': elem,
                            'text': data_value,
                            'selector': f'[{attr}]',
                            'method': 'data_attribute'
                        })
            
            # 3. Ищем элементы с определенными тегами, которые часто содержат цены
            price_tags = ['span', 'div', 'p', 'td', 'strong', 'b', 'i']
            for tag in price_tags:
                elements = soup.find_all(tag, text=re.compile(r'[\d\s.,]+(?:руб|р\.|₽|RUB|RUR)?', re.IGNORECASE))
                for elem in elements:
                    text = elem.get_text(strip=True)
                    if text and len(text) > 1:
                        all_elements.append({
                            'element': elem,
                            'text': text,
                            'selector': tag,
                            'method': 'tag_with_price'
                        })
            
            # 4. Ищем таблицы и строки таблиц
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        if text and len(text) > 1:
                            all_elements.append({
                                'element': cell,
                                'text': text,
                                'selector': 'table cell',
                                'method': 'table_cell'
                            })
            
            # Удаляем дубликаты (по тексту)
            unique_elements = []
            seen_texts = set()
            for elem in all_elements:
                if elem['text'] not in seen_texts:
                    seen_texts.add(elem['text'])
                    unique_elements.append(elem)
            
            logger.info(f"Найдено {len(unique_elements)} потенциальных элементов с ценами")
            return unique_elements
        except Exception as e:
            logger.error(f"Ошибка поиска потенциальных элементов с ценами: {e}")
            return []
        
    def universal_table_parsing_bestly(self, html, product_name, extracted_params=None):
        """
        Универсальный парсер для таблиц на bestly.ru.
        Теперь использует структуру таблицы для точного поиска.
        """
        try:
            logger.info(f"Используем универсальный парсер таблиц Bestly для: {product_name}")
            
            # Извлекаем параметры из названия товара
            parameters = self.extract_all_parameters_from_product_name(product_name)
            
            if not parameters or not parameters.get('thickness'):
                logger.warning(f"Не удалось определить толщину для товара: {product_name}")
                return None
            
            thickness = parameters['thickness']
            logger.info(f"Ищем цену для толщины: {thickness}мм")
            
            soup = BeautifulSoup(html, 'html.parser')
            
            # Ищем все таблицы на странице
            tables = soup.find_all('table')
            logger.info(f"Найдено таблиц: {len(tables)}")
            
            # Найдем таблицу с товарами
            for table_idx, table in enumerate(tables):
                rows = table.find_all('tr')
                
                # Ищем таблицу с товарами (должна быть достаточно большой)
                if len(rows) < 3:
                    continue
                
                # Пробуем найти заголовки таблицы
                first_row = rows[0]
                headers = [cell.get_text(strip=True).lower() for cell in first_row.find_all(['th', 'td'])]
                
                # Проверяем, есть ли в заголовках упоминание толщины
                has_thickness_header = any('толщин' in header for header in headers)
                
                if has_thickness_header:
                    logger.info(f"Найдена таблица #{table_idx+1} с заголовком толщины")
                    logger.info(f"Заголовки: {headers}")
                    
                    # Найдем индекс колонки с толщиной
                    thickness_col_idx = -1
                    for i, header in enumerate(headers):
                        if 'толщин' in header:
                            thickness_col_idx = i
                            break
                    
                    if thickness_col_idx == -1:
                        continue
                    
                    # Теперь ищем колонку с ценой
                    price_col_idx = -1
                    for i, header in enumerate(headers):
                        if 'цена' in header or '₽' in header:
                            price_col_idx = i
                            break
                    
                    if price_col_idx == -1:
                        # Если не нашли по "цена", ищем по другим признакам
                        for i, header in enumerate(headers):
                            if any(indicator in header for indicator in ['₽', 'руб', 'р.', 'стоимость']):
                                price_col_idx = i
                                break
                    
                    if price_col_idx == -1:
                        logger.warning(f"Не найдена колонка с ценой в таблице #{table_idx+1}")
                        continue
                    
                    logger.info(f"Индекс колонки толщины: {thickness_col_idx}, цены: {price_col_idx}")
                    
                    # Теперь ищем строку с нужной толщиной
                    for row_idx, row in enumerate(rows[1:], 1):  # Пропускаем заголовок
                        cells = row.find_all(['td', 'th'])
                        
                        if len(cells) <= max(thickness_col_idx, price_col_idx):
                            continue
                        
                        # Получаем значение толщины из соответствующей колонки
                        thickness_cell = cells[thickness_col_idx]
                        thickness_text = thickness_cell.get_text(strip=True)

                        # Проверяем, содержит ли ячейка нужную толщину
                        thickness_found = self._text_matches_thickness(
                            thickness_text, thickness, mode='exact', tolerance=0.01
                        )

                        if thickness_found:
                            logger.info(f"Найдена строка с толщиной {thickness}мм: '{thickness_text}'")
                            
                            # Получаем цену из соответствующей колонки
                            price_cell = cells[price_col_idx]
                            price_text = price_cell.get_text(strip=True)
                            price = self.extract_price_from_text(price_text)
                            
                            if price:
                                logger.info(f"Цена из колонки цены: '{price_text}' -> {price}")
                                return price
                            else:
                                # Если не нашли в колонке цены, пробуем найти цену во всей строке
                                row_text = row.get_text(strip=True)
                                price = self.extract_price_from_text(row_text)
                                if price:
                                    logger.info(f"Цена из всей строки: {price}")
                                    return price
            
            # Если не нашли через структуру таблицы, используем старый метод как запасной вариант
            logger.warning("Не удалось найти цену через структуру таблицы, используем альтернативный поиск...")
            return self.find_product_in_alternative_structures(html, product_name, parameters)
            
        except Exception as e:
            logger.error(f"Ошибка в универсальном парсере таблиц: {e}")
            traceback.print_exc()
            return None

    def extract_all_parameters_from_product_name(self, product_name):
        """
        Извлекает ВСЕ возможные параметры из названия товара.
        Возвращает словарь с параметрами.
        """
        parameters = {}
        
        try:
            if not product_name:
                return parameters
            
            clean_name = product_name.lower().strip()
            logger.info(f"Анализируем название: {clean_name}")
            
            # 1. Извлекаем толщину
            thickness = self.extract_thickness_from_product_name(product_name)
            if thickness:
                parameters['thickness'] = thickness
                parameters['thickness_mm'] = f"{thickness}мм"
                logger.info(f"Извлечена толщина: {thickness}")
            
            # 2. Извлекаем цвет
            color = self.extract_color_from_product_name(product_name)
            if color:
                parameters['color'] = color
            
            # 3. Извлекаем тип материала
            material_type = self.extract_material_type_from_product_name(product_name)
            if material_type:
                parameters['material_type'] = material_type
            
            # 4. Извлекаем размеры (если есть)
            dimensions = self.extract_dimensions_from_product_name(product_name)
            if dimensions:
                parameters['dimensions'] = dimensions
            
            # 5. Извлекаем бренд/марку
            brand = self.extract_brand_from_product_name(product_name)
            if brand:
                parameters['brand'] = brand
            
            # 6. Извлекаем тип продукта (сотовый, монолитный и т.д.)
            product_type = self.extract_product_type_from_product_name(product_name)
            if product_type:
                parameters['product_type'] = product_type
            
            logger.info(f"Извлечены параметры из '{product_name}': {parameters}")
            return parameters
            
        except Exception as e:
            logger.error(f"Ошибка извлечения параметров: {e}")
            return parameters

    def extract_material_type_from_product_name(self, product_name):
        """
        Извлекает тип материала из названия товара.
        Примеры: поликарбонат, оргстекло, пластик, акрил и т.д.
        """
        try:
            if not product_name:
                return None
            
            clean_name = product_name.lower().strip()

            for keyword, material in MATERIAL_TYPE_KEYWORDS.items():
                if keyword in clean_name:
                    return material

            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения типа материала: {e}")
            return None

    def extract_dimensions_from_product_name(self, product_name):
        """
        Извлекает размеры из названия товара.
        Примеры: 2100x12000, 2050х3050 и т.д.
        """
        try:
            if not product_name:
                return None
            
            clean_name = product_name.lower().strip()
            
            # Паттерн для размеров: цифры, разделитель x/х/×, цифры
            patterns = [
                r'(\d+)[хx×](\d+)',
                r'(\d+)\s*[хx×]\s*(\d+)',
                r'(\d+)[*x](\d+)'
            ]
            
            for pattern in patterns:
                match = re.search(pattern, clean_name)
                if match:
                    dimensions = f"{match.group(1)}x{match.group(2)}"
                    return dimensions
            
            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения размеров: {e}")
            return None

    def extract_brand_from_product_name(self, product_name):
        """
        Извлекает бренд/марку из названия товара.
        Примеры: POLYGAL, PLAZCRYL и т.д.
        """
        try:
            if not product_name:
                return None
            
            # Ищем слова в верхнем регистре (английские бренды)
            uppercase_words = re.findall(r'\b[A-Z]{2,}\b', product_name)
            if uppercase_words:
                # Берем первое слово из 2+ букв в верхнем регистре
                for word in uppercase_words:
                    if len(word) >= 3:  # Минимум 3 буквы
                        return word
            
            # Ищем известные бренды
            clean_name = product_name.lower()
            for brand in KNOWN_BRANDS:
                if brand in clean_name:
                    return brand
            
            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения бренда: {e}")
            return None

    def extract_product_type_from_product_name(self, product_name):
        """
        Извлекает тип продукта из названия.
        Примеры: сотовый, монолитный, листовой, профилированный и т.д.
        """
        try:
            if not product_name:
                return None
            
            clean_name = product_name.lower().strip()

            for keyword, product_type in PRODUCT_TYPE_KEYWORDS.items():
                if keyword in clean_name:
                    return product_type

            return None
        except Exception as e:
            logger.error(f"Ошибка извлечения типа продукта: {e}")
            return None

    def calculate_match_score(self, row_text, product_name, parameters):
        """
        Рассчитывает оценку соответствия строки параметрам товара.
        Теперь толщина является ОБЯЗАТЕЛЬНЫМ параметром.
        """
        try:
            score = 0
            row_text_lower = row_text.lower()
            
            # 1. ПРОВЕРКА ТОЛЩИНЫ - ОБЯЗАТЕЛЬНОЕ УСЛОВИЕ
            if parameters.get('thickness'):
                thickness = parameters['thickness']
                
                # Ищем толщину разными способами
                found_thickness = self._thickness_regex_match(row_text_lower, thickness, mode='boundary')

                # Также проверяем числовое совпадение, но ИСКЛЮЧАЯ числа в ценах
                if not found_thickness:
                    # Разделяем текст на слова и проверяем каждое слово отдельно
                    words = re.findall(r'\b\w+(?:[.,]\w+)?\b', row_text_lower)
                    
                    for word in words:
                        try:
                            # Проверяем, является ли слово числом
                            num = float(word.replace(',', '.'))
                            thickness_val = float(thickness)
                            
                            # Проверяем точное совпадение
                            if abs(num - thickness_val) < 0.1:
                                # Убедимся, что это не часть цены
                                # Цена обычно имеет формат "4 752.00" или "4752.00"
                                # Ищем контекст: если перед числом есть "₽", "руб", "р." или после числа есть "₽", "руб", "р."
                                word_index = row_text_lower.find(word)
                                context_before = row_text_lower[max(0, word_index-10):word_index]
                                context_after = row_text_lower[word_index+len(word):word_index+len(word)+10]
                                
                                # Если в контексте есть признаки цены, это не толщина
                                price_indicators = ['₽', 'руб', 'р.', 'цена', 'стоимость']
                                is_price_context = any(indicator in context_before or indicator in context_after 
                                                    for indicator in price_indicators)
                                
                                if not is_price_context:
                                    found_thickness = True
                                    break
                        except Exception:
                            continue
                
                # Если толщина не найдена - строка НЕ подходит
                if not found_thickness:
                    return 0  # Возвращаем 0, строка не будет рассматриваться
                
                # Если толщина найдена - даем высокий балл
                score += 50
                
                # Дополнительный балл за точное указание "мм"
                if re.search(f'{thickness}\\s*мм', row_text_lower, re.IGNORECASE):
                    score += 10
            
            # 2. Проверяем цвет (важный параметр для поликарбоната и оргстекла)
            if parameters.get('color'):
                color = parameters['color'].lower()
                if color in row_text_lower:
                    score += 20  # Высокий приоритет для цвета
            
            # 3. Проверяем размеры
            if parameters.get('dimensions'):
                dimensions = parameters['dimensions'].lower()
                if dimensions in row_text_lower:
                    score += 15
            
            # 4. Проверяем бренд
            if parameters.get('brand'):
                brand = parameters['brand'].lower()
                if brand in row_text_lower:
                    score += 10
            
            # 5. Проверяем тип материала
            if parameters.get('material_type'):
                material = parameters['material_type'].lower()
                if material in row_text_lower:
                    score += 5
            
            # 6. Проверяем тип продукта
            if parameters.get('product_type'):
                product_type = parameters['product_type'].lower()
                if product_type in row_text_lower:
                    score += 5
            
            return max(0, score)
            
        except Exception as e:
            logger.error(f"Ошибка расчета оценки: {e}")
            return 0
        
    def debug_row_info(self, row_text, parameters):
        """
        Выводит отладочную информацию о строке таблицы.
        """
        logger.info(f"Анализируем строку: {row_text[:100]}...")
        if parameters.get('thickness'):
            thickness = parameters['thickness']
            logger.info(f"Ищем толщину: {thickness}")
            
            # Ищем толщину разными способами
            for pattern in self._thickness_regex_patterns(thickness, mode='boundary'):
                match = re.search(pattern, row_text.lower())
                if match:
                    logger.info(f"Найдена толщина по паттерну '{pattern}': {match.group()}")

            # Ищем все числа в строке
            logger.info(f"Все числа в строке: {self._numbers_in_text(row_text)}")

    def extract_important_keywords(self, product_name):
        """
        Извлекает важные ключевые слова из названия товара.
        Исключает общие слова.
        """
        try:
            if not product_name:
                return []   
            
            words = re.findall(r'[а-яa-z0-9]{2,}', product_name.lower())

            # Фильтруем стоп-слова и короткие слова
            keywords = []
            for word in words:
                if (word not in KEYWORD_STOP_WORDS and
                    len(word) > 2 and
                    not word.isdigit() and
                    not (word.isdigit() and len(word) < 3)):  # Исключаем маленькие числа
                    keywords.append(word)
            
            return keywords
        except Exception as e:
            logger.error(f"Ошибка извлечения ключевых слов: {e}")
            return []

    def find_price_in_row(self, cells, row_text):
        """
        Находит цену в строке таблицы.
        Проверяет все ячейки и весь текст строки.
        """
        try:
            # Сначала ищем в ячейках, которые содержат явные признаки цены
            for cell in cells:
                cell_text = cell.get_text(strip=True)
                
                # Пропускаем пустые ячейки
                if not cell_text or len(cell_text) < 2:
                    continue
                
                # Проверяем, содержит ли ячейка признаки цены
                price_indicators = ['₽', 'руб', 'р.', '/л.', 'цена', 'стоимость']
                cell_text_lower = cell_text.lower()
                
                if any(indicator in cell_text_lower for indicator in price_indicators):
                    # Извлекаем цену из этой ячейки
                    price = self.extract_price_from_text(cell_text)
                    if price and self.is_reasonable_price(price, 'any'):
                        logger.info(f"Найдена цена в ячейке с признаком цены: {cell_text} -> {price}")
                        return price
            
            # Если не нашли по признакам, проверяем все ячейки
            for cell in cells:
                cell_text = cell.get_text(strip=True)
                
                # Пропускаем пустые ячейки
                if not cell_text or len(cell_text) < 2:
                    continue
                
                # Пропускаем ячейки, которые содержат только числа (возможно, это размеры или толщина)
                if re.match(r'^\d+$', cell_text):
                    continue
                
                # Пропускаем ячейки с размерами (содержат x или х)
                if re.search(r'\d+[хx×]\d+', cell_text):
                    continue
                
                # Извлекаем цену из ячейки
                price = self.extract_price_from_text(cell_text)
                if price and self.is_reasonable_price(price, 'any'):
                    logger.info(f"Найдена цена в ячейке: {cell_text} -> {price}")
                    return price
            
            # Если не нашли в ячейках, ищем во всем тексте строки
            price = self.extract_price_from_text(row_text)
            if price and self.is_reasonable_price(price, 'any'):
                logger.info(f"Найдена цена в тексте строки: {row_text[:100]}... -> {price}")
                return price
            
            return None
        except Exception as e:
            logger.error(f"Ошибка поиска цены в строке: {e}")
            return None

    def is_reasonable_price(self, price, product_name):
        """
        Проверяет, является ли цена разумной для данного товара.
        """
        try:
            # Базовые проверки для всех товаров
            if price <= 0:
                return False

            # Если имя товара не указано, используем общие проверки
            if not product_name or product_name == 'any':
                low, high = UNKNOWN_PRODUCT_PRICE_RANGE
                return low <= price <= high

            product_name_lower = product_name.lower()

            # Ищем первое правило, для которого встречаются ВСЕ его ключевые слова
            for rule in PRODUCT_PRICE_RANGES:
                if all(keyword in product_name_lower for keyword in rule['keywords']):
                    return rule['min'] <= price <= rule['max']

            # Общий диапазон для других товаров
            low, high = DEFAULT_PRICE_RANGE
            return low <= price <= high
        except Exception as e:
            logger.error(f"Ошибка проверки цены: {e}")
            low, high = DEFAULT_PRICE_RANGE
            return low <= price <= high  # Безопасный диапазон

    def find_product_in_alternative_structures(self, html, product_name, parameters):
        """
        Ищет товар в альтернативных структурах (не таблицы).
        Теперь использует более точный поиск по толщине.
        """
        try:
            logger.info(f"Ищем товар в альтернативных структурах: {product_name}")
            
            if not parameters.get('thickness'):
                return None
            
            thickness = parameters['thickness']
            soup = BeautifulSoup(html, 'html.parser')
            
            # 1. Ищем элементы, которые могут содержать информацию о товарах
            # (карточки товаров, блоки с характеристиками и т.д.)
            product_selectors = [
                '.product-item', '.catalog-item', '.item', '.goods-item',
                '.product-card', '.card', '.product', '.item-card',
                '.catalog-section-item', '.bx_catalog_item', '.item_block'
            ]
            
            for selector in product_selectors:
                elements = soup.select(selector)
                if not elements:
                    continue
                
                logger.info(f"Найдено {len(elements)} элементов по селектору '{selector}'")
                
                for element in elements:
                    element_text = element.get_text(strip=True)
                    
                    # Проверяем толщину в элементе
                    if self.check_thickness_in_text(element_text, thickness):
                        # Ищем цену в этом элементе
                        price = self.find_price_in_element(element)
                        if price and self.is_reasonable_price(price, product_name):
                            logger.info(f"Найдена цена в элементе: {price}")
                            return price
            
            # 2. Ищем все элементы с ценами
            price_elements = soup.find_all(text=re.compile(r'[\d\s.,]+(?:руб|р\.|₽)', re.IGNORECASE))
            
            for price_element in price_elements:
                price_text = str(price_element)
                price = self.extract_price_from_text(price_text)
                
                if price and self.is_reasonable_price(price, product_name):
                    # Проверяем контекст на наличие толщины
                    parent = price_element.parent
                    for _ in range(3):  # Проверяем 3 уровня родителей
                        if parent:
                            parent_text = parent.get_text(strip=True)
                            if self.check_thickness_in_text(parent_text, thickness):
                                logger.info(f"Найдена цена в контексте с толщиной {thickness}мм: {price}")
                                return price
                            parent = parent.parent
            
            return None
            
        except Exception as e:
            logger.error(f"Ошибка поиска в альтернативных структурах: {e}")
            return None

    def check_thickness_in_text(self, text, thickness):
        """
        Проверяет, содержит ли текст указанную толщину.
        """
        return self._text_matches_thickness(text, thickness, mode='boundary', tolerance=0.1)

    def find_price_in_element(self, element):
        """
        Находит цену в HTML-элементе.
        """
        try:
            # Ищем элементы с классами, содержащими price
            price_selectors = [
                '.price', '[class*="price"]', '[class*="Price"]',
                '.cost', '.value', '.amount', '.product-price',
                '.item-price', '.card-price', '.catalog-price'
            ]
            
            for selector in price_selectors:
                price_elements = element.select(selector)
                for price_element in price_elements:
                    price_text = price_element.get_text(strip=True)
                    price = self.extract_price_from_text(price_text)
                    if price:
                        return price
            
            # Если не нашли по селекторам, ищем во всем элементе
            element_text = element.get_text(strip=True)
            return self.extract_price_from_text(element_text)
        except Exception as e:
            logger.error(f"Ошибка поиска цены в элементе: {e}")
            return None

    def parse_orgsteklo_table_improved(self, html, product_name):
        """
        Улучшенный парсер для таблицы оргстекла на bestly.ru.
        Теперь ищет точное соответствие толщины в отдельном столбце.
        """
        try:
            logger.info(f"Используем улучшенный парсер для оргстекла: {product_name}")
            
            # Извлекаем толщину из названия товара
            thickness = self.extract_thickness_from_product_name(product_name)
            if not thickness:
                logger.warning(f"Не удалось определить толщину для товара: {product_name}")
                return None
            
            logger.info(f"Ищем цену для толщины: {thickness}мм")
            
            soup = BeautifulSoup(html, 'html.parser')
            
            # Ищем все таблицы на странице
            tables = soup.find_all('table')
            logger.info(f"Найдено таблиц: {len(tables)}")
            
            for table_idx, table in enumerate(tables):
                rows = table.find_all('tr')
                
                # Пропускаем таблицы без строк
                if len(rows) < 2:
                    continue
                
                # Ищем строки с толщиной и ценой
                for row_idx, row in enumerate(rows):
                    cells = row.find_all(['td', 'th'])
                    if len(cells) < 3:  # Нужно минимум 3 ячейки: толщина, размер, цена
                        continue
                    
                    # Пытаемся извлечь толщину из первой ячейки
                    thickness_cell_text = cells[0].get_text(strip=True)

                    # Проверяем, содержит ли ячейка нашу толщину
                    thickness_found = self._text_matches_thickness(
                        thickness_cell_text, thickness, mode='exact', tolerance=0.1
                    )

                    if thickness_found:
                        logger.info(f"Найдена строка с толщиной {thickness}мм в таблице {table_idx+1}, строка {row_idx+1}")
                        logger.info(f"Текст ячейки с толщиной: '{thickness_cell_text}'")
                        
                        # Теперь ищем цену в этой строке
                        # Обычно цена находится в последней ячейке или предпоследней
                        for cell_idx, cell in enumerate(cells):
                            cell_text = cell.get_text(strip=True)
                            
                            # Пропускаем ячейку с толщиной
                            if cell_idx == 0:
                                continue
                            
                            # Пропускаем ячейки, которые могут быть размерами
                            # Размеры обычно содержат "x" или "х" между цифрами
                            if re.search(r'\d{3,4}[хx×]\d{3,4}', cell_text):
                                continue
                            
                            # Пропускаем пустые ячейки
                            if not cell_text or len(cell_text) < 2:
                                continue
                            
                            # Извлекаем цену из ячейки
                            price = self.extract_price_from_text(cell_text)
                            if price and 1000 <= price <= 50000:  # Фильтруем по разумному диапазону
                                logger.info(f"Найдена цена в ячейке {cell_idx+1}: {price} (текст: '{cell_text}')")
                                return price
                        
                        # Если не нашли в отдельных ячейках, пробуем извлечь из всей строки
                        row_text = row.get_text(strip=True)
                        logger.info(f"Весь текст строки: '{row_text}'")
                        
                        # Извлекаем все числа из строки
                        all_numbers = re.findall(r'\d+(?:[.,]\d+)?', row_text)
                        logger.info(f"Все числа в строке: {all_numbers}")
                        
                        # Конвертируем в float
                        numbers_float = []
                        for num_str in all_numbers:
                            try:
                                num = float(num_str.replace(',', '.'))
                                numbers_float.append(num)
                            except Exception:
                                continue
                        
                        # Фильтруем числа
                        filtered_numbers = []
                        for num in numbers_float:
                            # Пропускаем толщину
                            if abs(num - float(thickness)) < 0.1:
                                continue
                            
                            # Пропускаем числа, которые могут быть размерами (1000-9999)
                            if 1000 <= num <= 9999:
                                # Проверяем, упоминается ли этот размер в названии товара
                                size_pattern = str(int(num))
                                if size_pattern in product_name.replace(' ', ''):
                                    continue
                            
                            # Пропускаем слишком маленькие числа
                            if num < 100:
                                continue
                            
                            filtered_numbers.append(num)
                        
                        logger.info(f"Отфильтрованные числа: {filtered_numbers}")
                        
                        if filtered_numbers:
                            # Берем максимальное число (обычно цена самая большая)
                            price = max(filtered_numbers)
                            logger.info(f"Выбрана цена: {price}")
                            return price
            
            # Если не нашли в таблицах, используем старый метод как fallback
            logger.warning("Не удалось найти цену в таблицах, используем старый метод...")
            
            # Старый метод поиска по всем элементам
            all_elements = self.find_all_potential_price_elements(html)
            
            if not all_elements:
                logger.warning("Не найдено потенциальных элементов с ценами")
                return None
            
            # Создаем словарь для хранения толщин и цен
            thickness_price_map = {}
            
            for elem in all_elements:
                text = elem['text']
                
                # Ищем толщину в тексте
                thickness_match = re.search(r'(\d+(?:[.,]\d+)?)\s*мм', text, re.IGNORECASE)
                if thickness_match:
                    found_thickness = thickness_match.group(1)
                    # Извлекаем ценю
                    price = self.extract_price_from_text(text)
                    if price:
                        # Преобразуем толщину в число
                        try:
                            thickness_val = float(found_thickness.replace(',', '.'))
                            thickness_price_map[thickness_val] = price
                            logger.info(f"Найдена толщина {thickness_val}мм с ценой {price}")
                        except Exception:
                            continue
            
            # Теперь ищем цену для нужной толщины
            target_thickness = float(thickness)
            
            # Ищем точное совпадение
            if target_thickness in thickness_price_map:
                logger.info(f"Найдена точная цена для толщины {target_thickness}мм: {thickness_price_map[target_thickness]}")
                return thickness_price_map[target_thickness]
            
            # Ищем ближайшую толщину
            if thickness_price_map:
                closest_thickness = min(thickness_price_map.keys(), key=lambda x: abs(x - target_thickness))
                logger.info(f"Найдена ближайшая толщина {closest_thickness}мм с ценой {thickness_price_map[closest_thickness]}")
                return thickness_price_map[closest_thickness]
            
            logger.warning(f"Не удалось найти цену для толщины {thickness}мм")
            return None
            
        except Exception as e:
            logger.error(f"Ошибка в улучшенном парсере оргстекла: {e}")
            traceback.print_exc()
            return None
    
    def parse_bestly_orgsteklo_table(self, html, product_name):
        """
        Новый парсер для таблицы оргстекла на bestly.ru.
        Ищет нужный товар по толщине и цвету в таблице.
        """
        try:
            logger.info(f"Используем новый парсер для таблицы оргстекла Bestly: {product_name}")
            
            # Извлекаем толщину из названия товара
            thickness = self.extract_thickness_from_product_name(product_name)
            if not thickness:
                logger.warning(f"Не удалось определить толщину для товара: {product_name}")
                return None
            
            # Извлекаем цвет из названия товара
            color = self.extract_color_from_product_name(product_name)
            if not color:
                logger.warning(f"Не удалось определить цвет для товара: {product_name}")
                # Если цвет не указан, ищем только по толщине
                color = None
            
            logger.info(f"Ищем цену для толщины: {thickness}мм, цвет: {color if color else 'любой'}")
            
            soup = BeautifulSoup(html, 'html.parser')
            
            # Находим все строки таблицы с товарами
            rows = []
            
            # Пробуем разные селекторы для строк таблицы
            table_selectors = [
                'table tbody tr',
                'table tr',
                '.table tbody tr',
                '.table tr',
                'tbody tr',
                '.catalog-table tr',
                '.price-table tr'
            ]
            
            for selector in table_selectors:
                rows = soup.select(selector)
                if rows and len(rows) > 1:  # Если нашли хотя бы 2 строки (заголовок + данные)
                    logger.info(f"Найдено {len(rows)} строк по селектору '{selector}'")
                    break
            
            if not rows:
                logger.warning("Не найдено строк таблицы на странице")
                return None
            
            found_rows = []
            
            for row_idx, row in enumerate(rows):
                # Извлекаем данные из ячеек этой строки
                cells = row.find_all(['td', 'th'])
                
                # Пропускаем строки без ячеек
                if not cells:
                    continue
                
                # Пропускаем строки с слишком маленьким количеством ячеек
                if len(cells) < 5:  # Минимум 5 ячеек: чекбокс, цвет, толщина, размер, цена
                    continue
                
                # Получаем текст из всех ячеек
                cell_texts = [cell.get_text(strip=True) for cell in cells]
                
                # Пропускаем строки, где первая ячейка содержит текст (это может быть заголовок)
                if cell_texts[0] and len(cell_texts[0]) > 10:
                    continue
                
                # Ищем толщину в ячейках - ТОЧНОЕ СОВПАДЕНИЕ В ЯЧЕЙКЕ ТОЛЩИНЫ
                # В таблице bestly.ru структура обычно: чекбокс, цвет, толщина, размер, цена
                # Проверяем только ячейку с индексом 2 (третья ячейка - толщина)
                thickness_found = False

                if len(cells) > 2:
                    thickness_cell_text = cell_texts[2]  # Третья ячейка должна быть толщиной
                    thickness_found = self._text_matches_thickness(
                        thickness_cell_text, thickness, mode='exact', tolerance=0.01
                    )
                
                # Ищем цвет, если он указан
                color_found = False if color else True  # Если цвет не указан, считаем, что он найден
                
                if color and len(cells) > 1:  # Вторая ячейка должна быть цветом
                    color_cell_text = cell_texts[1]
                    if color in color_cell_text.lower():
                        color_found = True
                
                # Если нашли и толщину, и (если нужен) цвет
                if thickness_found and color_found:
                    logger.info(f"Найдена подходящая строка #{row_idx+1}")
                    logger.info(f"Текст ячеек: {cell_texts}")
                    
                    # Ищем цену в этой строке (обычно 4-я или 5-я ячейка)
                    for i, cell_text in enumerate(cell_texts):
                        # Пропускаем первые 3 ячейки (чекбокс, цвет, толщина)
                        if i < 3:
                            continue
                        
                        # Пропускаем ячейки с размерами (содержат x или х)
                        if re.search(r'\d{3,4}[хx×]\d{3,4}', cell_text):
                            continue
                        
                        # Пропускаем пустые ячейки
                        if not cell_text or len(cell_text) < 2:
                            continue
                        
                        # Извлекаем цену из ячейки
                        price = self.extract_price_from_text(cell_text)
                        if price and 1000 <= price <= 50000:  # Фильтруем по разумному диапазону
                            logger.info(f"Найдена цена в строке #{row_idx+1}, ячейка {i+1}: {price} (текст: '{cell_text}')")
                            found_rows.append({
                                'row_idx': row_idx,
                                'cell_texts': cell_texts,
                                'price': price,
                                'thickness_cell': cell_texts[2] if len(cell_texts) > 2 else '',
                                'color_cell': cell_texts[1] if len(cell_texts) > 1 else '',
                                'price_cell': cell_text,
                                'thickness_match': thickness_found,
                                'color_match': color_found
                            })
                            break
                    
                    # Если не нашли в отдельных ячейках, пробуем извлечь из всей строки
                    if not found_rows:
                        row_text = row.get_text(strip=True)
                        price = self.extract_price_from_text(row_text)
                        if price and 1000 <= price <= 50000:
                            logger.info(f"Найдена цена из всего текста строки #{row_idx+1}: {price}")
                            found_rows.append({
                                'row_idx': row_idx,
                                'cell_texts': cell_texts,
                                'price': price,
                                'thickness_cell': cell_texts[2] if len(cell_texts) > 2 else '',
                                'color_cell': cell_texts[1] if len(cell_texts) > 1 else '',
                                'price_cell': row_text[:100],
                                'thickness_match': thickness_found,
                                'color_match': color_found
                            })
            
            if found_rows:
                # Сортируем найденные строки по точности совпадения
                def sort_key(row_info):
                    score = 0
                    
                    # Высший приоритет: точное совпадение толщины в ячейке толщины
                    thickness_cell = row_info['thickness_cell']
                    if re.search(f'^{thickness}\\s*мм$', thickness_cell, re.IGNORECASE):
                        score += 100
                    elif re.search(f'\\b{thickness}\\s*мм\\b', thickness_cell, re.IGNORECASE):
                        score += 90
                    elif re.search(f'\\b{thickness}\\b', thickness_cell):
                        score += 80
                    
                    # Проверяем числовое совпадение
                    try:
                        numbers = re.findall(r'\d+(?:[.,]\d+)?', thickness_cell)
                        for num_str in numbers:
                            num = float(num_str.replace(',', '.'))
                            if abs(num - float(thickness)) < 0.01:
                                score += 70
                                break
                    except Exception:
                        pass
                    
                    # Приоритет для правильного цвета
                    if color and row_info['color_match']:
                        score += 50
                    
                    # Приоритет для более высокой строки (ближе к началу таблицы)
                    score -= row_info['row_idx'] * 0.1
                    
                    return score
                
                found_rows.sort(key=sort_key, reverse=True)
                
                # Выбираем первую строку после сортировки
                best_row = found_rows[0]
                logger.info(f"Выбрана строка #{best_row['row_idx']+1} с ценой: {best_row['price']}")
                logger.info(f"Толщина в ячейке: '{best_row['thickness_cell']}'")
                logger.info(f"Цвет в ячейке: '{best_row['color_cell']}'")
                logger.info("Причина выбора: лучший match по толщине и цвету")
                return best_row['price']
            else:
                logger.warning(f"Не найден товар с толщиной {thickness}мм и цветом {color if color else 'любым'}")
                return None
                
        except Exception as e:
            logger.error(f"Ошибка в новом парсере таблицы оргстекла: {e}")
            traceback.print_exc()
            return None
        
    def find_product_by_name_in_context(self, element, product_name):
        """Ищет название продукта в контексте элемента и его родителей"""
        try:
            if not product_name or not element:
                return False
            
            # Очищаем название продукта для поиска
            clean_product_name = product_name.lower().strip()
            
            # Проверяем текст самого элемента
            element_text = element.get_text(strip=True).lower()
            if clean_product_name in element_text:
                return True
            
            # Проверяем родителей (до 3 уровней вверх)
            parent = element.parent
            for i in range(3):
                if parent and parent.name not in ['html', 'body']:
                    parent_text = parent.get_text(strip=True).lower()
                    if clean_product_name in parent_text:
                        return True
                    parent = parent.parent
                else:
                    break
            
            # Ищем название в соседних элементах
            try:
                # Ищем в предыдущих и следующих элементах
                siblings = list(element.parent.children) if element.parent else []
                if siblings:
                    element_index = siblings.index(element)
                    
                    # Проверяем предыдущие элементы
                    for i in range(max(0, element_index - 3), element_index):
                        if i < len(siblings):
                            sibling_text = str(siblings[i]).lower()
                            if clean_product_name in sibling_text:
                                return True
                    
                    # Проверяем следующие элементы
                    for i in range(element_index + 1, min(len(siblings), element_index + 4)):
                        sibling_text = str(siblings[i]).lower()
                        if clean_product_name in sibling_text:
                            return True
            except Exception:
                pass
            
            return False
            
        except Exception as e:
            logger.error(f"Ошибка поиска названия в контексте: {e}")
            return False
    
    def find_price_with_selector_and_name(self, html, selector, product_name):
        """Поиск цены по селектору с фильтрацией по названию товара"""
        try:
            soup = BeautifulSoup(html, 'html.parser')

            if not selector or not selector.strip():
                return None

            elements = soup.select(selector.strip())
            if not elements:
                return None

            def extract_reasonable_price(element):
                # Отбрасываем значения, явно не похожие на цену (например,
                # data-price="1" от виджета количества товара, а не самой
                # цены — так на bestly.ru находилась "цена" 1 руб.)
                price_text = element.get_text(strip=True)
                extracted_price = self.extract_price_from_text(price_text)
                if extracted_price and self.is_reasonable_price(extracted_price, product_name):
                    return extracted_price
                return None

            # Если элементов несколько, ищем тот, который связан с названием товара
            if len(elements) > 1 and product_name:
                logger.info(f"Найдено {len(elements)} элементов по селектору '{selector}', ищем по названию товара...")

                # Сначала ищем элемент, в контексте которого есть название товара
                for element in elements:
                    if self.find_product_by_name_in_context(element, product_name):
                        price = extract_reasonable_price(element)
                        if price:
                            logger.info(f"Найдена цена по селектору с названием товара: {price}")
                            return price

                # Если не нашли по названию — ищем первую цену, прошедшую
                # проверку на разумность (а не просто первый элемент)
                logger.info("Не найдено цен по названию товара, ищем первую разумную цену среди найденных элементов")
                for element in elements:
                    price = extract_reasonable_price(element)
                    if price:
                        logger.info(f"Возвращаем первую разумную найденную цену: {price}")
                        return price
                return None

            # Если элемент один или название не указано — берём первый элемент
            # с разумной ценой (может быть не самый первый в списке, если
            # селектор случайно зацепил что-то ещё)
            for element in elements:
                price = extract_reasonable_price(element)
                if price:
                    logger.info(f"Найдена цена по селектору '{selector}': {price}")
                    return price

            return None

        except Exception as e:
            logger.error(f"Ошибка поиска цены по селектору с названием: {e}")
            return None
    
    def find_price_with_selenium_by_name(self, product_name, price_selector=None):
        """
        Поиск цены с помощью Selenium, используя название товара для нахождения карточки товара.
        """
        if not self.driver:
            logger.error("Selenium драйвер не инициализирован")
            return None

        try:
            # Разбиваем название товара на ключевые слова
            keywords = [word.strip().lower() for word in product_name.split() if len(word.strip()) > 2]
            if not keywords:
                return None

            # Подготавливаем фразы для поиска: полное название, первые 5 слов, первые 3 слова
            search_phrases = [product_name.lower()]
            if len(keywords) > 5:
                search_phrases.append(' '.join(keywords[:5]))
            if len(keywords) > 3:
                search_phrases.append(' '.join(keywords[:3]))

            # Пробуем найти по фразам
            for phrase in search_phrases:
                try:
                    # Ищем элемент, содержащий фразу (без учета регистра через translate)
                    element = self.driver.find_element(
                        By.XPATH, 
                        f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ', 'abcdefghijklmnopqrstuvwxyzабвгдеёжзийклмнопрстуфхцчшщъыьэюя'), '{phrase}')]"
                    )
                    # Находим карточку товара, поднимаясь по дереву
                    card = element
                    for _ in range(5):  # Максимум 5 уровней вверх
                        try:
                            card_class = card.get_attribute('class').lower() if card.get_attribute('class') else ''
                            if any(cls in card_class for cls in ['item', 'product', 'card', 'catalog']):
                                break
                        except Exception:
                            pass
                        card = card.find_element(By.XPATH, "./..")
                    
                    # Ищем цену в карточке
                    if price_selector:
                        try:
                            price_element = card.find_element(By.CSS_SELECTOR, price_selector)
                        except Exception:
                            price_element = None
                    else:
                        price_element = None
                        # Используем селекторы из конфигурации bestly.ru
                        for sel in self.site_configs.get('bestly.ru', {}).get('price_selectors', []):
                            try:
                                price_element = card.find_element(By.CSS_SELECTOR, sel)
                                break
                            except Exception:
                                continue
                    
                    if price_element:
                        price_text = price_element.text
                        price = self.extract_price_from_text(price_text)
                        if price:
                            return price
                except Exception as e:
                    logger.debug(f"Не удалось найти по фразе '{phrase}': {e}")
                    continue

            # Если по фразам не нашли, ищем по ключевым словам в карточках
            card_selectors = self.site_configs.get('bestly.ru', {}).get('card_selectors', [])
            for card_selector in card_selectors:
                try:
                    cards = self.driver.find_elements(By.CSS_SELECTOR, card_selector)
                    for card in cards:
                        try:
                            card_text = card.text.lower()
                            # Проверяем, содержатся ли первые 3 ключевых слова в тексте карточки
                            if all(keyword in card_text for keyword in keywords[:3]):
                                # Ищем цену в карточке
                                if price_selector:
                                    try:
                                        price_element = card.find_element(By.CSS_SELECTOR, price_selector)
                                    except Exception:
                                        price_element = None
                                else:
                                    price_element = None
                                    for sel in self.site_configs.get('bestly.ru', {}).get('price_selectors', []):
                                        try:
                                            price_element = card.find_element(By.CSS_SELECTOR, sel)
                                            break
                                        except Exception:
                                            continue
                                if price_element:
                                    price_text = price_element.text
                                    price = self.extract_price_from_text(price_text)
                                    if price:
                                        return price
                        except Exception:
                            continue
                except Exception:
                    continue

            return None
        except Exception as e:
            logger.error(f"Ошибка в Selenium контекстном поиске для товара '{product_name}': {e}")
            return None
    
    def find_price_and_name_on_page(self, html, url, selector=None, product_name=None):
        """Поиск цены и названия товара на странице"""
        prices_found = []
        found_name = None
        
        # Сначала пробуем указанный селектор с фильтрацией по названию
        if selector and selector.strip():
            price = self.find_price_with_selector_and_name(html, selector, product_name)
            if price:
                # Ищем название товара на странице
                found_name = self.extract_product_name_from_page(html)
                prices_found.append({
                    'price': price, 
                    'method': 'specified_selector_with_name', 
                    'selector': selector,
                    'product_name': found_name
                })
        
        # Для bestly.ru используем специальную логику
        if 'bestly.ru' in url:
            # Сначала пробуем новый парсер для таблицы оргстекла
            if 'org_steklo' in url.lower() or 'стекло' in (product_name or '').lower():
                logger.info("Пробуем новый парсер для таблицы оргстекла Bestly...")
                table_price = self.parse_bestly_orgsteklo_table(html, product_name or '')
                if table_price:
                    found_name = self.extract_product_name_from_page(html)
                    prices_found.append({
                        'price': table_price, 
                        'method': 'bestly_table_parser', 
                        'selector': 'table parsing',
                        'product_name': found_name
                    })
                    logger.info(f"Найдена цена через новый парсер таблицы: {table_price}")
            
            # Также пробуем специальные селекторы
            bestly_selectors = self.site_configs.get('bestly.ru', {}).get('price_selectors', [])
            logger.info(f"Используем специальные селекторы для Bestly: {len(bestly_selectors)} селекторов")
            
            for sel in bestly_selectors:
                price = self.find_price_with_selector_and_name(html, sel, product_name)
                if price:
                    # Ищем название товара на странице
                    found_name = self.extract_product_name_from_page(html)
                    prices_found.append({
                        'price': price, 
                        'method': 'bestly_specific_with_name', 
                        'selector': sel,
                        'product_name': found_name
                    })
                    logger.info(f"Найдена цена по селектору Bestly '{sel}' с названием: {price}")
                    break
        
        # Пробуем найти все элементы с числами, которые могут быть ценами
        soup = BeautifulSoup(html, 'html.parser')
        
        # Ищем элементы, содержащие цифры и возможные индикаторы цены
        price_candidates = []
        
        # Ищем по текстовым элементам
        all_texts = soup.find_all(text=re.compile(r'[\d\s.,]+(?:руб|р\.|₽|RUB|RUR)?', re.IGNORECASE))
        for text in all_texts:
            price = self.extract_price_from_text(str(text))
            # Отсеиваем неправдоподобные значения (например, "404" из текста
            # "Страница не найдена (404 Not Found)" на несуществующей странице)
            if price and self.is_reasonable_price(price, product_name):
                # Получаем родительский элемент для определения селектор
                parent = text.parent
                if parent:
                    # Пытаемся определить селектор
                    selector = self._get_selector_for_element(parent)
                    # Проверяем, есть ли название товара в контексте
                    has_product_name = self.find_product_by_name_in_context(parent, product_name) if product_name else True
                    
                    # Извлекаем возможное название товара
                    candidate_name = None
                    if has_product_name and product_name:
                        candidate_name = product_name
                    else:
                        candidate_name = self.extract_product_name_from_page(html)
                    
                    price_candidates.append({
                        'price': price,
                        'text': str(text)[:100],
                        'selector': selector,
                        'has_product_name': has_product_name,
                        'product_name': candidate_name
                    })
        
        # Ищем элементы с атрибутами данных, содержащими цены
        for elem in soup.find_all(attrs={"data-price": True}):
            data_price = elem.get('data-price')
            price = self.extract_price_from_text(str(data_price))
            # Отсеиваем неправдоподобные значения (например, data-price="1"
            # у виджета количества товара, а не саму цену)
            if price and self.is_reasonable_price(price, product_name):
                selector = self._get_selector_for_element(elem)
                has_product_name = self.find_product_by_name_in_context(elem, product_name) if product_name else True
                
                candidate_name = None
                if has_product_name and product_name:
                    candidate_name = product_name
                else:
                    candidate_name = self.extract_product_name_from_page(html)
                
                price_candidates.append({
                    'price': price,
                    'text': data_price,
                    'selector': selector or '[data-price]',
                    'has_product_name': has_product_name,
                    'product_name': candidate_name
                })
        
        # Ищем элементы с классами, содержащими price, cost, value и т.д.
        price_patterns = ['price', 'cost', 'value', 'amount', 'стоимость', 'цена']
        for pattern in price_patterns:
            elements = soup.find_all(class_=re.compile(pattern, re.IGNORECASE))
            for elem in elements:
                text = elem.get_text(strip=True)
                price = self.extract_price_from_text(text)
                if price and self.is_reasonable_price(price, product_name):
                    selector = self._get_selector_for_element(elem)
                    has_product_name = self.find_product_by_name_in_context(elem, product_name) if product_name else True

                    candidate_name = None
                    if has_product_name and product_name:
                        candidate_name = product_name
                    else:
                        candidate_name = self.extract_product_name_from_page(html)

                    price_candidates.append({
                        'price': price,
                        'text': text[:100],
                        'selector': selector or f'[class*="{pattern}"]',
                        'has_product_name': has_product_name,
                        'product_name': candidate_name
                    })
        
        # Убираем дубликаты по цене
        unique_prices = []
        seen_prices = set()
        for candidate in price_candidates:
            if candidate['price'] not in seen_prices:
                seen_prices.add(candidate['price'])
                unique_prices.append(candidate)
        
        # Сортируем по вероятности (предполагаем, что правильная цена обычно среднего диапазона)
        # И приоритезируем те, где есть название товара
        def price_score(candidate):
            score = 0
            price = candidate['price']
            
            # Приоритет для цен, найденных в контексте с названием товара
            if candidate.get('has_product_name', False):
                score += 20
            
            # Цены от 100 до 10000 считаем более вероятными
            if 100 <= price <= 10000:
                score += 10
            
            # Если в селекторе есть price или cost
            if candidate['selector'] and any(word in candidate['selector'].lower() for word in ['price', 'cost', 'value']):
                score += 5
            
            # Если в тексте есть руб или р.
            if candidate['text'] and any(word in candidate['text'].lower() for word in ['руб', 'р.', '₽']):
                score += 3
            
            return score
        
        unique_prices.sort(key=lambda x: price_score(x), reverse=True)
        
        # Добавляем лучшие кандидаты
        for candidate in unique_prices[:5]:  # Берем топ-5 кандидатов
            prices_found.append({
                'price': candidate['price'],
                'method': 'auto_detected',
                'selector': candidate['selector'],
                'has_product_name': candidate.get('has_product_name', False),
                'product_name': candidate.get('product_name')
            })
        
        # Находим лучшее название товара
        if prices_found and not found_name:
            # Используем название из первого найденного кандидата
            for price_info in prices_found:
                if price_info.get('product_name'):
                    found_name = price_info['product_name']
                    break
        
        return prices_found, found_name
    
    def _get_selector_for_element(self, element):
        """Генерирует селектор для элемента"""
        try:
            # Пытаемся создать селектор на основе класса
            classes = element.get('class', [])
            if classes:
                # Берем первый непустой класс
                for cls in classes:
                    if cls and cls.strip():
                        return f'.{cls.strip()}'
            
            # Пытаемся использовать id
            elem_id = element.get('id')
            if elem_id:
                return f'#{elem_id}'
            
            # Используем тег
            tag = element.name
            return tag
            
        except Exception:
            return None
    
    def get_with_requests(self, url, headers=None):
        """
        Получение страницы с помощью requests.

        Побочный эффект: сохраняет в self.last_fetched_url финальный URL
        после всех редиректов (site.ru может сделать 301 со старого адреса
        товара на новый, если его ЧПУ-слаг пересобрался) и в
        self.last_status_code — HTTP-статус ответа. Вызывающий код может
        использовать last_fetched_url, чтобы обновить сохраненную ссылку
        на товар и не держать её "протухшей" вручную.
        """
        self.last_fetched_url = None
        self.last_status_code = None
        try:
            if headers is None:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                    'Accept-Encoding': 'gzip, deflate, br',
                    'Connection': 'keep-alive',
                }

            response = requests.get(
                url,
                headers=headers,
                timeout=60,  # Увеличиваем таймаут для Bestly
                verify=True,
                allow_redirects=True
            )
            self.last_status_code = response.status_code

            if response.status_code != 200:
                logger.warning(f"Страница вернула статус {response.status_code}: {url}")
                return None

            # Если сайт сделал редирект на другой адрес (например, ЧПУ-ссылка
            # товара изменилась) — запоминаем финальный URL для самолечения
            resolved_url = strip_tracking_params(response.url)
            if resolved_url and resolved_url != url:
                logger.info(f"Ссылка была перенаправлена: {url} -> {resolved_url}")
            self.last_fetched_url = resolved_url

            # Определяем кодировку
            if response.encoding is None or response.encoding == 'ISO-8859-1':
                response.encoding = 'utf-8'

            return response.text

        except requests.exceptions.Timeout:
            logger.error(f"Таймаут при запросе к {url}")
            return None
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка запроса к {url}: {e}")
            return None
        except Exception as e:
            logger.error(f"Неожиданная ошибка при запросе к {url}: {e}")
            return None
    
    def get_with_selenium(self, url, wait_time=None):
        """Получение страницы с помощью Selenium с улучшенной поддержкой Bestly"""
        if not self.driver:
            if not self.init_selenium_driver():
                logger.error("Не удалось инициализировать Selenium драйвер")
                return None
        
        try:
            logger.info(f"Selenium: загрузка {url}")
            
            # Устанавливаем таймаут загрузки страницы
            self.driver.set_page_load_timeout(120)  # Увеличиваем таймаут для Bestly
            
            # Для bestly.ru очищаем cookies и кэш перед загрузкой
            if 'bestly.ru' in url:
                try:
                    self.driver.delete_all_cookies()
                    logger.info("Cookies очищены для Bestly")
                except Exception:
                    pass
            
            # Загружаем страницу
            self.driver.get(url)
            
            # Настраиваем время ожидания в зависимости от сайта
            if 'bestly.ru' in url:
                wait_time = wait_time or 20  # Увеличиваем ожидание для Bestly
                logger.info(f"Ожидание загрузки Bestly: до {wait_time} секунд...")

                # Вместо слепого time.sleep(wait_time) ждем появления хотя бы
                # одного элемента, похожего на цену/таблицу — обычно это
                # происходит быстрее верхней границы, а на медленных
                # страницах мы всё равно ждем не дольше, чем раньше.
                try:
                    WebDriverWait(self.driver, wait_time).until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, 'table, [class*="price"], [class*="Price"], [data-price]')
                        )
                    )
                except TimeoutException:
                    logger.warning(f"Не дождались признаков цены за {wait_time}с, продолжаем как есть")

                # Прокручиваем несколько раз для загрузки динамического контента
                scroll_attempts = 8  # Увеличиваем количество прокруток
                found_prices_during_scroll = False
                for i in range(scroll_attempts):
                    try:
                        # Прокручиваем страницу
                        scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        current_scroll = (scroll_height / scroll_attempts) * (i + 1)
                        self.driver.execute_script(f"window.scrollTo(0, {current_scroll});")
                        logger.info(f"Прокрутка {i+1}/{scroll_attempts} до позиции {current_scroll}")
                        time.sleep(4)  # Увеличиваем ожидание после прокрутки

                        # Пробуем найти элементы с ценами
                        price_elements = self.driver.find_elements(By.CSS_SELECTOR, '[class*="price"], [class*="Price"], [data-price], .price, .Price')
                        if price_elements and i >= 3:  # Если нашли цены после 4-й прокрутки
                            logger.info(f"Найдено {len(price_elements)} элементов с ценами после прокрутки")
                            found_prices_during_scroll = True
                            break

                    except Exception as e:
                        logger.warning(f"Ошибка при прокрутке: {e}")
                        continue

                # Дополнительное ожидание для динамического контента.
                # Если цены уже найдены во время прокрутки — не ждем полные
                # 8 секунд впустую, короткой паузы достаточно.
                time.sleep(2 if found_prices_during_scroll else 8)

                # Пробуем кликнуть на кнопки "Показать еще" или аналогичные
                try:
                    show_more_buttons = self.driver.find_elements(By.XPATH, "//button[contains(text(), 'Показать') or contains(text(), 'Еще') or contains(text(), 'Загрузить') or contains(text(), 'Показать еще')]")
                    for button in show_more_buttons[:3]:  # Пробуем первые 3 кнопки
                        try:
                            button.click()
                            time.sleep(5)
                            logger.info("Кликнут на кнопку 'Показать еще'")
                        except Exception:
                            pass
                except Exception:
                    pass
                    
            else:
                # Для других сайтов используем стандартное ожидание
                wait_time = wait_time or 10
                try:
                    WebDriverWait(self.driver, wait_time).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                except TimeoutException:
                    logger.warning(f"Страница не сообщила о готовности за {wait_time}с, продолжаем как есть")
                self.driver.execute_script("window.scrollTo(0, 500);")
                time.sleep(2)
            
            # Получаем HTML
            page_source = self.driver.page_source
            
            # Проверяем, что страница загрузилась
            if not page_source or len(page_source) < 1000:
                logger.warning(f"Страница {url} загрузилась с пустым или маленьким HTML ({len(page_source) if page_source else 0} символов)")
                return None
            
            logger.info(f"Страница успешно загружена, размер HTML: {len(page_source)} символов")
            return page_source
                
        except TimeoutException:
            logger.error(f"Таймаут при загрузке страницы {url}")
            # Пробуем получить хотя бы частично загруженную страницу
            try:
                page_source = self.driver.page_source
                if page_source and len(page_source) > 1000:
                    logger.warning(f"Используем частично загруженную страницу {url}")
                    return page_source
                else:
                    return None
            except Exception:
                return None
        except Exception as e:
            logger.error(f"Ошибка Selenium при загрузке {url}: {str(e)[:200]}")
            # Закрываем драйвер при ошибке
            self.close_selenium_driver()
            return None
    
    def find_and_select_selector_for_product(self):
        """Найти и выбрать селектор для товара"""
        try:
            if not self.load_excel_data():
                print("Не удалось загрузить данные из Excel")
                return
            
            print(f"\nВсего товаров: {len(self.df)}")
            idx_str = input("Введите номер товара для выбора селектора (начиная с 1): ").strip()
            
            if not idx_str.isdigit():
                print("Неверный номер товара")
                return
            
            idx = int(idx_str) - 1
            if idx < 0 or idx >= len(self.df):
                print("Неверный номер товара")
                return
            
            row = self.df.iloc[idx]
            name = safe_str(row.get('Название', f'Товар {idx+1}'))
            url = strip_tracking_params(safe_str(row.get('URL', '')).strip())
            
            if not url:
                print("URL не указан для этого товара")
                return
            
            print(f"\nТовар #{idx+1}: {name}")
            print(f"URL: {url}")
            
            # Показываем сохраненный выбор пользователя, если есть
            saved_selection = self.get_user_selection(url)
            if saved_selection:
                print("\nСохраненный выбор пользователя:")
                print(f"  Селектор: {saved_selection['selector']}")
                print(f"  Текст элемента: {saved_selection['text'][:100]}...")
                print(f"  Цена: {saved_selection['price']}")
                print(f"  Время сохранения: {saved_selection['timestamp']}")
            
            # Загружаем страницу
            print("\nЗагружаем страницу...")
            domain = self.extract_domain(url)
            html = None
            
            if domain in self.site_configs and self.site_configs[domain].get('method') == 'selenium':
                html = self.get_with_selenium(url)
            else:
                headers = self.site_configs.get(domain, {}).get('headers', {})
                html = self.get_with_requests(url, headers)
            
            if not html:
                print("Не удалось загрузить страницу")
                return
            
            # Находим все похожие элементы
            print("\nИщем похожие элементы на странице...")
            
            # Для bestly.ru используем специальные селекторы
            if 'bestly.ru' in domain:
                selectors = self.site_configs.get('bestly.ru', {}).get('price_selectors', [])
            else:
                # Используем общие селекторы
                selectors = [
                    '.price', '.product-price', '.woocommerce-Price-amount',
                    '.product-item-price', '.regular-price', '.current-price',
                    '.product-card__price', '.catalog-item-price',
                    'span.price', 'div.price', 'p.price',
                    '[data-price]', '[itemprop="price"]',
                    '.product-card-price', '.item-price', '.price-value',
                    '.product_price', '.card-price', '.catalog-price'
                ]
            
            all_elements = self.find_similar_elements_by_selector(html, selectors)
            
            if not all_elements:
                print("Не найдено похожих элементов на странице")
                
                # Пробуем найти все потенциальные элементы с ценами
                print("Пробуем найти все потенциальные элементы с ценами...")
                all_elements = self.find_all_potential_price_elements(html)
                
                if not all_elements:
                    print("Не найдено ни одного элемента с ценами")
                    return
            
            print(f"\nНайдено {len(all_elements)} похожих элементов:")
            print("-" * 100)
            
            # Выводим список элементов
            for i, elem in enumerate(all_elements):
                price = self.extract_price_from_text(elem['text'])
                price_str = f"Цена: {price}" if price else "Цена не найдена"
                print(f"{i+1}. [{elem['selector']}]")
                print(f"   Текст: {elem['text'][:150]}...")
                print(f"   {price_str}")
                print("-" * 100)
            
            # Предлагаем выбрать элемент
            choice_str = input("\nВыберите номер правильного элемента (или 0 для отмены): ").strip()
            
            if not choice_str.isdigit():
                print("Отмена выбора")
                return
            
            choice = int(choice_str) - 1
            
            if choice == -1:
                print("Отмена выбора")
                return
            
            if choice < 0 or choice >= len(all_elements):
                print("Неверный номер элемента")
                return
            
            # Получаем выбранный элемент
            selected_elem = all_elements[choice]
            selected_price = self.extract_price_from_text(selected_elem['text'])
            
            if not selected_price:
                print("В выбранном элементе не удалось найти цену")
                return
            
            # Сохраняем выбор пользователя
            print("\nВыбран элемент:")
            print(f"  Селектор: {selected_elem['selector']}")
            print(f"  Текст: {selected_elem['text'][:100]}...")
            print(f"  Цена: {selected_price}")
            
            # Предлагаем сохранить выбор
            save_choice = input("\nСохранить этот выбор? (да/нет): ").strip().lower()
            
            if save_choice in ['да', 'yes', 'y']:
                # Сохраняем выбор пользователя
                if self.save_user_selection(url, selected_elem['selector'], selected_elem['text'], selected_price):
                    print("✓ Выбор сохранен")
                    
                    # Также обновляем селектор в Excel
                    try:
                        wb = load_workbook(self.excel_file)
                        if self.sheet_name in wb.sheetnames:
                            ws = wb[self.sheet_name]
                            
                            # Находим колонку "Селектор"
                            sel_col = None
                            for col in range(1, ws.max_column + 1):
                                header = ws.cell(row=1, column=col).value
                                if header and 'селектор' in str(header).lower():
                                    sel_col = col
                                    break
                            
                            if sel_col:
                                ws.cell(row=idx+2, column=sel_col).value = selected_elem['selector']
                                wb.save(self.excel_file)
                                print("✓ Селектор сохранен в Excel")
                            else:
                                print("⚠ Не найдена колонка 'Селектор' в Excel")
                    except Exception as e:
                        print(f"Ошибка сохранения в Excel: {e}")
                else:
                    print("Ошибка сохранения выбора")
            else:
                print("Выбор не сохранен")
            
            input("\nНажмите Enter для возврата в меню...")
            
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()
            input("\nНажмите Enter для возврата в меню...")

    def test_universal_parser(self):
        """Тестирование универсального парсера таблиц"""
        try:
            if not self.load_excel_data():
                print("Не удалось загрузить данные из Excel")
                return
            
            print(f"\nВсего товаров: {len(self.df)}")
            print("Найдены товары с bestly.ru:")
            
            bestly_products = []
            for idx in range(len(self.df)):
                row = self.df.iloc[idx]
                name = safe_str(row.get('Название', f'Товар {idx+1}'))
                url = strip_tracking_params(safe_str(row.get('URL', '')).strip())
                
                if 'bestly.ru' in url:
                    bestly_products.append((idx, name, url))
            
            if not bestly_products:
                print("Не найдено товаров с bestly.ru")
                return
            
            for i, (idx, name, url) in enumerate(bestly_products):
                print(f"{i+1}. #{idx+1}: {name}")
            
            choice = input("\nВыберите товар для тестирования (номер из списка): ").strip()
            if not choice.isdigit():
                print("Неверный выбор")
                return
            
            choice_idx = int(choice) - 1
            if choice_idx < 0 or choice_idx >= len(bestly_products):
                print("Неверный номер")
                return
            
            idx, name, url = bestly_products[choice_idx]
            
            print(f"\nТестирование товара #{idx+1}: {name}")
            print(f"URL: {url}")
            
            # Извлекаем параметры для проверки
            parameters = self.extract_all_parameters_from_product_name(name)
            print(f"Извлеченные параметры: {parameters}")
            
            # Загружаем страницу
            print("\nЗагружаем страницу...")
            html = self.get_with_selenium(url)
            
            if not html:
                print("Не удалось загрузить страницу")
                return
            
            # Используем универсальный парсер
            print("\nИспользуем универсальный парсер таблиц...")
            price = self.universal_table_parsing_bestly(html, name)
            
            if price:
                print(f"\n✓ Найдена цена: {price} руб.")
                
                # Показываем подробности
                soup = BeautifulSoup(html, 'html.parser')
                tables = soup.find_all('table')
                print(f"Найдено таблиц на странице: {len(tables)}")
                
                for i, table in enumerate(tables[:3]):  # Показываем первые 3 таблицы
                    rows = table.find_all('tr')
                    print(f"\nТаблица #{i+1}: {len(rows)} строк")
                    
                    # Показываем первые 5 строк
                    for j, row in enumerate(rows[:6]):
                        cells = row.find_all(['td', 'th'])
                        cell_texts = [cell.get_text(strip=True)[:30] for cell in cells]
                        print(f"  Строка {j}: {cell_texts}")
            else:
                print("\n✗ Цена не найдена")
                
                # Пробуем альтернативный поиск
                print("Пробуем альтернативный поиск...")
                alt_price = self.find_product_in_alternative_structures(html, name, parameters)
                if alt_price:
                    print(f"✓ Найдена цена альтернативным методом: {alt_price} руб.")
            
            input("\nНажмите Enter для возврата в меню...")
            
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()
    
    def test_selector_for_product(self):
        """Тестирование селектора для товара"""
        try:
            if not self.load_excel_data():
                print("Не удалось загрузить данные из Excel")
                return
            
            print(f"\nВсего товаров: {len(self.df)}")
            idx_str = input("Введите номер товара для тестирования селектора (начиная с 1): ").strip()
            
            if not idx_str.isdigit():
                print("Неверный номер товара")
                return
            
            idx = int(idx_str) - 1
            if idx < 0 or idx >= len(self.df):
                print("Неверный номер товара")
                return
            
            row = self.df.iloc[idx]
            name = safe_str(row.get('Название', f'Товар {idx+1}'))
            url = strip_tracking_params(safe_str(row.get('URL', '')).strip())
            selector = safe_str(row.get('Селектор', '')).strip()
            
            if not url:
                print("URL не указан для этого товара")
                return
            
            if not selector:
                print("Селектор не указан для этого товара")
                return
            
            print(f"\nТовар #{idx+1}: {name}")
            print(f"URL: {url}")
            print(f"Селектор для тестирования: {selector}")
            
            # Загружаем страницу
            print("\nЗагружаем страницу...")
            domain = self.extract_domain(url)
            html = None
            
            if domain in self.site_configs and self.site_configs[domain].get('method') == 'selenium':
                html = self.get_with_selenium(url)
            else:
                headers = self.site_configs.get(domain, {}).get('headers', {})
                html = self.get_with_requests(url, headers)
            
            if not html:
                print("Не удалось загрузить страницу")
                return
            
            # Тестируем селектор
            print(f"\nТестируем селектор: {selector}")
            
            soup = BeautifulSoup(html, 'html.parser')
            elements = soup.select(selector.strip())
            
            if not elements:
                print("✗ Не найдено элементов по этому селектору")
            else:
                print(f"✓ Найдено {len(elements)} элементов по этому селектору")
                
                # Показываем все найденные элементы
                for i, element in enumerate(elements):
                    text = element.get_text(strip=True)
                    price = self.extract_price_from_text(text)
                    
                    print(f"\nЭлемент {i+1}:")
                    print(f"  Текст: {text[:150]}...")
                    print(f"  HTML: {str(element)[:200]}...")
                    if price:
                        print(f"  Цена: {price}")
                    else:
                        print("  Цена: не найдена в тексте")
            
            input("\nНажмите Enter для возврата в меню...")
            
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()
            input("\nНажмите Enter для возврата в меню...")
    
    def test_selector_multiple_prices(self, html, selector, product_name=None):
        """Тестирование селектора и поиск всех цен по нему"""
        try:
            soup = BeautifulSoup(html, 'html.parser')
            
            if not selector or not selector.strip():
                return []
            
            elements = soup.select(selector.strip())
            results = []
            
            for i, element in enumerate(elements):
                text = element.get_text(strip=True)
                price = self.extract_price_from_text(text)
                
                if price:
                    results.append({
                        'index': i,
                        'text': text[:200],
                        'price': price,
                        'html': str(element)[:300]
                    })
            
            return results
        except Exception as e:
            logger.error(f"Ошибка тестирования селектора: {e}")
            return []
    
    def save_selector_to_excel(self, idx, selector, price, name, characteristic=None):
        """Сохранение селектора в Excel"""
        try:
            wb = load_workbook(self.excel_file)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
            else:
                ws = wb[self.sheet_name]
            
            # Находим колонки
            sel_col = None
            price_col = None
            name_col = None
            char_col = None
            
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_str = str(header).lower()
                    if 'селектор' in header_str:
                        sel_col = col
                    elif 'цена' in header_str and 'за м2' not in header_str:
                        price_col = col
                    elif 'название' in header_str:
                        name_col = col
                    elif 'характеристика' in header_str:
                        char_col = col
            
            # Если нет колонки для селектора, создаем
            if sel_col is None:
                sel_col = ws.max_column + 1
                ws.cell(row=1, column=sel_col).value = "Селектор"
            
            # Если нет колонки для характеристики, создаем
            if char_col is None:
                char_col = ws.max_column + 1
                ws.cell(row=1, column=char_col).value = "Характеристика"
            
            row_idx = idx + 2
            
            # Сохраняем селектор
            ws.cell(row=row_idx, column=sel_col).value = selector
            
            # Сохраняем цену, если нашли
            if price and price_col:
                ws.cell(row=row_idx, column=price_col).value = price
            
            # Сохраняем характеристику, если передана
            if characteristic and char_col:
                ws.cell(row=row_idx, column=char_col).value = characteristic
            
            wb.save(self.excel_file)
            print(f"✓ Селектор для товара '{name}' сохранен в Excel")
            
        except Exception as e:
            print(f"Ошибка сохранения селектора: {e}")
    
    def parse_single_product(self, index, row):
        """Парсинг одного товара с сохранением выбранного селектора и поиском по названию"""
        try:
            # Получаем данные из строки
            name = safe_str(row.get('Название', f'Товар {index+1}'))
            url = strip_tracking_params(safe_str(row.get('URL', '')).strip())
            selector = safe_str(row.get('Селектор', '')).strip()
            characteristic = safe_str(row.get('Характеристика', '')).strip()
            
            logger.info(f"\n{'='*60}")
            logger.info(f"Парсинг товара #{index+1}: {name}")
            logger.info(f"URL: {url}")
            logger.info(f"Текущий селектор: {selector}")
            logger.info(f"Текущая характеристика: {characteristic}")
            logger.info(f"Режим округления: {self.rounding_mode}")
            
            if not url or url == 'nan' or url.lower() == 'none':
                logger.warning("Пропуск: URL не указан")
                return {
                    'index': index,
                    'name': name,
                    'url': '',
                    'price': None,
                    'characteristic': characteristic,
                    'selector_used': selector,
                    'best_found_selector': '',
                    'status': 'URL не указан',
                    'rounding_mode': self.rounding_mode,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            
            # Определяем домен ДО использования
            domain = self.extract_domain(url)
            
            # Проверяем, есть ли сохраненный выбор пользователя для этого URL
            saved_selection = self.get_user_selection(url)
            if saved_selection:
                logger.info("Найден сохраненный выбор пользователя для этого товара")
                logger.info(f"Сохраненный селектор: {saved_selection['selector']}")
                logger.info(f"Сохраненная цена: {saved_selection['price']}")
                
                # Используем сохраненный выбор
                selector = saved_selection['selector']
            
            # Определяем метод парсинга
            method = 'requests'  # По умолчанию
            
            # Для bestly.ru всегда используем Selenium
            if 'bestly.ru' in domain:
                method = 'selenium'
                logger.info("Используем Selenium для Bestly")
            
            # Получаем HTML
            html = None
            
            if method == 'selenium':
                html = self.get_with_selenium(url)
                if not html:
                    logger.warning(f"Selenium не смог загрузить страницу {url}, пробуем requests...")
                    method = 'requests'
            
            if method == 'requests' or not html:
                headers = self.site_configs.get(domain, {}).get('headers', {})
                html = self.get_with_requests(url, headers)

                # Сайт мог сделать редирект со старой ссылки на новую (например,
                # у товара пересобрался ЧПУ-адрес) — "самолечим" ссылку прямо
                # здесь, чтобы result['url'] попал в Excel уже актуальным и не
                # приходилось вручную обновлять её при каждом изменении.
                if html and self.last_fetched_url and self.last_fetched_url != url:
                    logger.info(f"Ссылка товара изменилась (редирект сайта): {url} -> {self.last_fetched_url}")
                    url = self.last_fetched_url
                    domain = self.extract_domain(url)

            if not html:
                was_404 = (self.last_status_code == 404)
                logger.warning(f"Ссылка не открылась, пробуем найти товар в каталоге: {url}")
                recovered_url, recovered_html = self.try_recover_and_refetch(url, name, domain)
                if recovered_url and recovered_html:
                    url = recovered_url
                    domain = self.extract_domain(url)
                    html = recovered_html
                else:
                    if was_404:
                        status = 'Ссылка больше не работает (404), похожий товар в каталоге не найден — нужно найти новый адрес вручную'
                    else:
                        status = 'Не удалось загрузить страницу'
                    return {
                        'index': index,
                        'name': name,
                        'url': url,
                        'price': None,
                        'characteristic': characteristic,
                        'selector_used': selector,
                        'best_found_selector': '',
                        'status': status,
                        'rounding_mode': self.rounding_mode,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }

            # "Мягкий" 404: сайт отдал HTTP 200, но содержимое — страница
            # вида "товар не найден" (см. looks_like_error_page). Без этой
            # проверки число вроде "404" из текста заглушки могло быть
            # ошибочно принято за цену товара.
            if self.looks_like_error_page(html):
                logger.warning(f"Страница похожа на заглушку 'не найдено' (HTTP 200), пробуем найти товар в каталоге: {url}")
                recovered_url, recovered_html = self.try_recover_and_refetch(url, name, domain)
                if recovered_url and recovered_html:
                    url = recovered_url
                    domain = self.extract_domain(url)
                    html = recovered_html
                else:
                    return {
                        'index': index,
                        'name': name,
                        'url': url,
                        'price': None,
                        'characteristic': characteristic,
                        'selector_used': selector,
                        'best_found_selector': '',
                        'status': 'Ссылка больше не работает (страница-заглушка «не найдено»), похожий товар в каталоге не найден — нужно найти новый адрес вручную',
                        'rounding_mode': self.rounding_mode,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }

            # ВАЖНО: Для Bestly используем УНИВЕРСАЛЬНЫЙ парсер таблиц
            is_bestly_page = 'bestly.ru' in domain
            
            # Если это bestly.ru и есть признаки табличного представления
            if is_bestly_page and html:
                # Проверяем, есть ли таблицы на странице
                soup = BeautifulSoup(html, 'html.parser')
                tables = soup.find_all('table')
                
                if tables or 'catalog' in url or 'товар' in name.lower():
                    logger.info(f"Обнаружена страница Bestly с табличной структурой: {name}")
                    
                    # ИСПРАВЛЕНИЕ: Используем правильный метод - universal_table_parsing_bestly
                    universal_price = self.universal_table_parsing_bestly(html, name)
                    
                    if universal_price:
                        # Успешно нашли цену
                        best_price = universal_price
                        best_method = 'universal_table_parser'
                        status = "Успешно (универсальный парсер таблиц Bestly)"
                        best_found_selector = 'universal_table_parsing'
                        
                        logger.info(f"✓ Найдена цена через универсальный парсер таблиц: {best_price}")
                        
                        # Ищем название товара на странице
                        found_product_name = self.extract_product_name_from_page(html)
                        if found_product_name:
                            characteristic = found_product_name
                        
                        result = {
                            'index': index,
                            'name': name,
                            'url': url,
                            'price': best_price,
                            'characteristic': characteristic,
                            'selector_used': selector,
                            'best_found_selector': best_found_selector,
                            'status': status,
                            'rounding_mode': self.rounding_mode,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        self.print_result(result)
                        return result
                    else:
                        logger.warning("Универсальный парсер не нашел цену, продолжаем стандартный поиск")
            
            # Стандартный поиск
            best_price = None
            best_characteristic = characteristic  # Начинаем с текущей характеристики
            best_selector = selector  # Начинаем с текущего селектора
            best_found_selector = ''
            best_method = 'not_found'
            status = 'Цена не найдена'
            found_product_name = None
            
            # Сначала пробуем использовать указанный селектор с фильтрацией по названию
            if selector and selector.strip():
                logger.info(f"Пробуем использовать селектор с названием: {selector}")

                # Используем функцию с фильтрацией по названию
                price = self.find_price_with_selector_and_name(html, selector, name)

                if price:
                    # Ищем название товара на странице
                    found_product_name = self.extract_product_name_from_page(html)
                    if found_product_name:
                        best_characteristic = found_product_name

                    best_price = price
                    best_selector = selector
                    best_found_selector = selector
                    best_method = 'specified_selector_with_name'
                    status = f"Успешно (указанный селектор с названием: {selector})"
                    logger.info(f"✓ Цена найдена по указанному селектору с названием: {best_price} руб.")
                else:
                    logger.info("✗ По указанному селектору с названием цена не найдена")
            else:
                logger.info("Селектор не указан, ищем любыми способами с названием...")

            # Если цена ещё не найдена (указанный селектор не сработал или не
            # был задан вовсе) — пробуем остальные способы РОВНО ОДИН РАЗ.
            #
            # Раньше этот блок был по ошибке продублирован и вложен так, что:
            #  1) если селектор не был указан вообще, этот код не выполнялся
            #     ни разу, и переменная result ниже оставалась неприсвоенной,
            #     из-за чего парсинг товара падал с UnboundLocalError на
            #     КАЖДОМ товаре без селектора (а он не указан почти всегда,
            #     пока пользователь не выберет его вручную через меню);
            #  2) если селектор был указан, но не сработал, а альтернативные
            #     селекторы Bestly находили цену — код всё равно проваливался
            #     в скопированный ниже блок и, не находя ничего через
            #     find_price_and_name_on_page(selector=None) во второй раз,
            #     обнулял уже найденную best_price обратно в None.
            # Оба случая тихо превращали найденную цену в "Цена не найдена".
            if not best_price:
                if 'bestly.ru' in domain:
                    logger.info("Для Bestly ищем специальные селекторы...")
                    bestly_selectors = self.site_configs.get('bestly.ru', {}).get('price_selectors', [])
                    for sel in bestly_selectors:
                        price = self.find_price_with_selector_and_name(html, sel, name)
                        if price:
                            found_product_name = self.extract_product_name_from_page(html)
                            if found_product_name:
                                best_characteristic = found_product_name

                            best_price = price
                            best_selector = sel
                            best_found_selector = sel
                            best_method = 'bestly_specific_with_name'
                            status = f"Успешно (специальный селектор Bestly: {sel})"
                            logger.info(f"✓ Цена найдена по специальному селектору Bestly '{sel}' с названием: {best_price} руб.")
                            break

                if not best_price:
                    logger.info("Ищем цену любыми способами с названием...")
                    price_results, found_name = self.find_price_and_name_on_page(html, url, selector=None, product_name=name)

                    if price_results:
                        # Берем первую найденную цену (уже отсортированы по наличию названия)
                        best_result = price_results[0]
                        best_price = best_result['price']
                        best_selector = best_result.get('selector', '')
                        best_found_selector = best_result.get('selector', '')
                        best_method = best_result['method']
                        status = f"Успешно ({best_method})"

                        # Обновляем характеристику, если нашли название
                        if found_name:
                            best_characteristic = found_name
                        elif best_result.get('product_name'):
                            best_characteristic = best_result.get('product_name')

                        logger.info(f"Найдено {len(price_results)} вариантов цен:")
                        for i, price_option in enumerate(price_results, 1):
                            name_flag = " [с названием]" if price_option.get('has_product_name', False) else ""
                            logger.info(f"  {i}. {price_option['price']} руб. (метод: {price_option['method']}, селектор: {price_option.get('selector', 'не указан')}){name_flag}")
                    else:
                        status = f"Цена не найдена (селектор: {selector})" if selector else 'Цена не найдена'

            # Fallback: для сайтов, которые используют Selenium и имеют флаг use_selenium_context_search
            if domain in self.site_configs and self.site_configs[domain].get('use_selenium_context_search', False):
                if not best_price and self.driver:
                    logger.info(f"Пробуем Selenium контекстный поиск для {name}")
                    selenium_price = self.find_price_with_selenium_by_name(name, selector)
                    if selenium_price:
                        best_price = selenium_price
                        best_method = 'selenium_context'
                        status = "Успешно (Selenium контекстный поиск)"
                        logger.info(f"Найдена цена через Selenium контекстный поиск: {best_price}")

            # Формируем результат
            result = {
                'index': index,
                'name': name,
                'url': url,
                'price': best_price,
                'characteristic': best_characteristic,  # Сохраняем найденную характеристику
                'selector_used': selector,  # Сохраняем исходный селектор
                'best_found_selector': best_found_selector,  # Найденный селектор
                'status': status,
                'rounding_mode': self.rounding_mode,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            # Выводим в консоль
            self.print_result(result)

            return result
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге товара {index+1}: {e}")
            traceback.print_exc()
            
            # В случае ошибки закрываем драйвер, если он есть
            self.close_selenium_driver()
            
            return {
                'index': index,
                'name': name if 'name' in locals() else f'Товар {index+1}',
                'url': url if 'url' in locals() else '',
                'price': None,
                'characteristic': characteristic if 'characteristic' in locals() else '',
                'selector_used': selector if 'selector' in locals() else '',
                'best_found_selector': '',
                'status': f'Ошибка: {str(e)[:100]}',
                'rounding_mode': self.rounding_mode,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
    
    def print_result(self, result):
        """Вывод результата в консоль"""
        print(f"\n{'='*60}")
        print(f"ТОВАР #{result['index'] + 1}")
        print(f"Название: {result['name'][:50]}...")
        print(f"URL: {result['url'][:80]}...")
        print(f"Использованный селектор: {result.get('selector_used', 'Не указан')}")
        print(f"Найденный селектор: {result.get('best_found_selector', 'Не найден')}")
        print(f"Характеристика: {result.get('characteristic', 'Не найдена')}")
        print(f"Режим округления: {result.get('rounding_mode', 'ceil')}")
        print(f"{'='*60}")
        
        if result['price'] is not None:
            print(f"✓ НАЙДЕНА ЦЕНА: {result['price']} руб.")
            rounding_mode = result.get('rounding_mode', 'ceil')
            if rounding_mode == 'ceil':
                print("  (округлено в большую сторону)")
            elif rounding_mode == 'floor':
                print("  (округлено в меньшую сторону)")
            elif rounding_mode == 'nearest':
                print("  (округлено до ближайшего целого)")
            elif rounding_mode == 'nearest_decimal':
                print("  (округлено до 2 знаков)")
            elif rounding_mode == 'no_decimal':
                print("  (округлено до целого)")
        else:
            print("✗ ЦЕНА: Не найдена")
        
        print(f"Статус: {result['status']}")
        print(f"Время: {result['timestamp']}")
        print(f"{'='*60}")
    
    def parse_all_products(self, start_from=0, limit=None):
        """Парсинг всех товаров"""
        if not self.load_excel_data():
            logger.error("Не удалось загрузить данные из Excel")
            return False
        
        self.results = []
        
        try:
            # Определяем диапазон для парсинга
            end_idx = len(self.df)
            if limit:
                end_idx = min(start_from + limit, len(self.df))
            
            print(f"\n{'='*60}")
            print("НАЧАЛО ПАРСИНГА")
            print(f"Всего товаров: {len(self.df)}")
            print(f"Начинаем с позиции: {start_from + 1}")
            print(f"Будет обработано: {end_idx - start_from}")
            print(f"Режим округления: {self.rounding_mode}")
            print("ВАЖНО: Для Bestly используется Selenium с улучшенной поддержкой")
            print("ВАЖНО: Для оргстекла на bestly.ru используется НОВЫЙ парсер таблиц")
            print("ПРИМЕЧАНИЕ: Используются сохраненные выборы пользователя")
            print("ПРИМЕЧАНИЕ: Цены ищутся по селектору и названию товара")
            print("ПРИМЕЧАНИЕ: Найденные характеристики сохраняются в столбец 'Характеристика'")
            print(f"{'='*60}")
            
            # Счетчики для статистики
            bestly_count = 0
            other_count = 0
            user_selection_used = 0
            table_parser_used = 0
            
            # Парсим товары
            for idx in range(start_from, end_idx):
                if idx >= len(self.df):
                    break
                
                row = self.df.iloc[idx]
                url = strip_tracking_params(safe_str(row.get('URL', '')).strip())
                
                # Определяем тип сайта
                domain = self.extract_domain(url)
                
                # Проверяем, есть ли сохраненный выбор пользователя
                saved_selection = self.get_user_selection(url)
                if saved_selection:
                    user_selection_used += 1
                    logger.info(f"Используем сохраненный выбор пользователя для товара {idx+1}")
                
                # Закрываем драйвер перед каждым bestly.ru, чтобы избежать накопления кэша
                if 'bestly.ru' in domain and self.driver:
                    self.close_selenium_driver()
                    bestly_count += 1
                elif domain:
                    other_count += 1
                
                # Парсим товар
                result = self.parse_single_product(idx, row)
                self.results.append(result)
                
                # Считаем использование нового парсера таблиц
                if 'bestly_table_parser' in result.get('status', ''):
                    table_parser_used += 1
                
                # Задержка между запросами (разная для разных сайтов)
                if idx < end_idx - 1:
                    if 'bestly.ru' in domain:
                        time.sleep(8)  # Большая задержка для Bestly
                    else:
                        time.sleep(3)  # Стандартная задержка
            
            print("\nСтатистика по типам сайтов:")
            print(f"  Bestly.ru: {bestly_count} товаров")
            print(f"  Другие сайты: {other_count} товаров")
            print(f"  Использовано сохраненных выборов: {user_selection_used}")
            print(f"  Использован новый парсер таблиц: {table_parser_used}")
            
            # Сохраняем результаты
            self.save_results_to_excel()
            
            # Генерируем отчет
            self.generate_report()
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге всех товаров: {e}")
            traceback.print_exc()
            return False
        finally:
            self.close_selenium_driver()
    
    def save_results_to_excel(self):
        """Сохранение результатов обратно в Excel"""
        try:
            # Открываем файл для записи
            wb = load_workbook(self.excel_file)
            
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
            else:
                ws = wb[self.sheet_name]
            
            # Определяем индексы колонок
            col_indices = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    col_name = str(cell_value).strip().lower()
                    
                    if 'название' in col_name:
                        col_indices['Название'] = col
                    elif 'цена' in col_name and 'за м2' not in col_name:
                        col_indices['Цена'] = col
                    elif 'url' in col_name or 'ссылка' in col_name:
                        col_indices['URL'] = col
                    elif 'селектор' in col_name:
                        col_indices['Селектор'] = col
                    elif 'характеристика' in col_name:
                        col_indices['Характеристика'] = col
                    elif 'дата' in col_name and 'обнов' in col_name:
                        col_indices['Дата обновления'] = col
            
            # Если колонки "Характеристика" нет, добавляем ее
            if 'Характеристика' not in col_indices:
                new_col_idx = ws.max_column + 1
                ws.cell(row=1, column=new_col_idx).value = "Характеристика"
                col_indices['Характеристика'] = new_col_idx
            
            # Обновляем данные
            for result in self.results:
                row_idx = result['index'] + 2  # +1 для заголовка, +1 для смещения
                
                if 'Цена' in col_indices:
                    ws.cell(row=row_idx, column=col_indices['Цена']).value = result['price']
                
                if 'Характеристика' in col_indices:
                    ws.cell(row=row_idx, column=col_indices['Характеристика']).value = result.get('characteristic', '')
                
                if 'Дата обновления' in col_indices:
                    ws.cell(row=row_idx, column=col_indices['Дата обновления']).value = result['timestamp']
                
                # Важно: НЕ обновляем селектор, если он уже есть
                if 'Селектор' in col_indices:
                    current_selector = ws.cell(row=row_idx, column=col_indices['Селектор']).value
                    # Обновляем селектор только если он был пустой
                    if not current_selector and result.get('best_found_selector'):
                        ws.cell(row=row_idx, column=col_indices['Селектор']).value = result['best_found_selector']

                # Самолечение ссылки: если сайт сделал редирект на новый адрес
                # товара (см. get_with_requests/parse_single_product), или из
                # ссылки убрали трекинг-параметры (ysclid и т.п.) — записываем
                # актуальный URL обратно в Excel, чтобы не редактировать его
                # руками при каждом изменении. Пустой result['url'] (например,
                # статус "URL не указан") сюда не попадает — ячейку не трогаем.
                if 'URL' in col_indices and result.get('url'):
                    current_url = ws.cell(row=row_idx, column=col_indices['URL']).value
                    new_url = result['url']
                    if current_url != new_url:
                        ws.cell(row=row_idx, column=col_indices['URL']).value = new_url
                        logger.info(f"Обновлена ссылка в Excel (строка {row_idx}): {current_url} -> {new_url}")

            # Сохраняем файл
            wb.save(self.excel_file)
            
            # ПРИМЕНЯЕМ ФОРМАТИРОВАНИЕ
            self.apply_formatting(wb, ws)
            
            logger.info(f"Результаты сохранены в файл: {self.excel_file}, лист: {self.sheet_name}")
            logger.info("Характеристики сохранены в столбец 'Характеристика'")
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка сохранения результатов: {e}")
            traceback.print_exc()
            return False
    
    def apply_formatting(self, wb, ws):
        """Применение форматирования к листу"""
        try:
            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Стиль для столбца Цена
            price_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            price_font = Font(color="000000", size=11)
            price_alignment = Alignment(horizontal="right", vertical="center")
            price_number_format = '#,##0.00'
            
            # Стиль для найденных цен (зеленый фон)
            price_found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            price_found_font = Font(color="006100", bold=True, size=11)
            
            # Стиль для отсутствующих цен (красный фон)
            price_missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            price_missing_font = Font(color="9C0006", bold=True, size=11)
            
            # Стиль для столбца Характеристика
            char_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            char_font = Font(color="000000", size=11)
            char_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Стиль для столбца Дата обновления
            date_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            date_font = Font(color="000000", size=11)
            date_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Форматируем заголовки
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Определяем индексы столбцов
            price_col_idx = None
            char_col_idx = None
            date_col_idx = None
            
            for col in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col).value
                if header_value:
                    header_str = str(header_value).lower()
                    
                    if 'цена' in header_str and 'за м2' not in header_str:
                        price_col_idx = col
                    elif 'характеристика' in header_str:
                        char_col_idx = col
                    elif 'дата' in header_str and 'обнов' in header_str:
                        date_col_idx = col
            
            # Форматируем данные
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    
                    # Форматирование столбца Цена
                    if col == price_col_idx:
                        cell.alignment = price_alignment
                        cell.number_format = price_number_format
                        
                        # Если цена найдена - зеленый фон, если нет - красный
                        if cell.value is not None and cell.value != '':
                            try:
                                # Преобразуем значение в число
                                if isinstance(cell.value, str):
                                    # Убираем пробелы и заменяем запятые на точки
                                    price_str = cell.value.replace(' ', '').replace(',', '.')
                                    price_val = float(price_str)
                                else:
                                    price_val = float(cell.value)
                                
                                # Устанавливаем форматированное значение
                                cell.value = price_val
                                cell.fill = price_found_fill
                                cell.font = price_found_font
                            except (ValueError, TypeError):
                                # Если не число, оставляем как есть с красным фоном
                                cell.fill = price_missing_fill
                                cell.font = price_missing_font
                        else:
                            # Пустая ячейка - красный фон
                            cell.fill = price_missing_fill
                            cell.font = price_missing_font
                    
                    # Форматирование столбца Характеристика
                    elif col == char_col_idx:
                        cell.fill = char_fill
                        cell.font = char_font
                        cell.alignment = char_alignment
                    
                    # Форматирование столбца Дата обновления
                    elif col == date_col_idx:
                        cell.fill = date_fill
                        cell.font = date_font
                        cell.alignment = date_alignment
                        
                        # Если дата есть, форматируем ее
                        if cell.value:
                            try:
                                # Если это datetime объект, преобразуем в строку
                                if hasattr(cell.value, 'strftime'):
                                    cell.value = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                                elif isinstance(cell.value, str):
                                    # Пытаемся разобрать строку с датой
                                    from datetime import datetime as dt
                                    # Пробуем разные форматы
                                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y']:
                                        try:
                                            date_obj = dt.strptime(cell.value, fmt)
                                            cell.value = date_obj.strftime('%Y-%m-%d %H:%M:%S')
                                            break
                                        except ValueError:
                                            continue
                            except Exception:
                                # Если не удалось преобразовать, оставляем как есть
                                pass
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass
                
                # Для столбца Цена устанавливаем фиксированную ширину
                if price_col_idx and ws.cell(row=1, column=price_col_idx).column_letter == column_letter:
                    adjusted_width = 15
                # Для столбца Характеристика устанавливаем большую ширину
                elif char_col_idx and ws.cell(row=1, column=char_col_idx).column_letter == column_letter:
                    adjusted_width = 40
                # Для столбца Дата устанавливаем фиксированную ширину
                elif date_col_idx and ws.cell(row=1, column=date_col_idx).column_letter == column_letter:
                    adjusted_width = 20
                else:
                    adjusted_width = min(max_length + 2, 50)
                
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            wb.save(self.excel_file)
            logger.info("Форматирование применено")
            
        except Exception as e:
            logger.error(f"Ошибка форматирования: {e}")
            traceback.print_exc()
    
    def generate_report(self):
        """Генерация отчета о парсинге"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_file = f"report_parsing_{timestamp}.txt"
            
            total = len(self.results)
            successful = sum(1 for r in self.results if r['price'] is not None)
            failed = total - successful
            
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write(f"{'='*70}\n")
                f.write("ОТЧЕТ ПАРСИНГА ЦЕН (СЕЛЕКТОРЫ СОХРАНЕНЫ, ПОИСК ПО НАЗВАНИЮ)\n")
                f.write(f"{'='*70}\n")
                f.write(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Файл: {self.excel_file}\n")
                f.write(f"Лист: {self.sheet_name}\n")
                f.write(f"Обработано: {total}\n")
                f.write(f"Режим округления: {self.rounding_mode}\n")
                f.write("Новый парсер для таблицы оргстекла Bestly: ВКЛЮЧЕН\n")
                f.write(f"Использовано сохраненных выборов: {len(self.user_selections)}\n")
                f.write(f"{'='*70}\n\n")
                
                f.write("СТАТИСТИКА:\n")
                f.write(f"  Успешно: {successful}\n")
                f.write(f"  Не найдено: {failed}\n")
                if total > 0:
                    f.write(f"  Процент успеха: {(successful/total*100):.1f}%\n\n")
                
                # Статистика по новому парсеру таблиц
                table_parser_results = [r for r in self.results if 'новый парсер таблицы' in r.get('status', '').lower()]
                if table_parser_results:
                    f.write("НОВЫЙ ПАРСЕР ТАБЛИЦЫ BESTLY.RU:\n")
                    f.write(f"  Обработано новым парсером: {len(table_parser_results)}\n")
                    table_parser_success = sum(1 for r in table_parser_results if r['price'] is not None)
                    f.write(f"  Успешно найдено: {table_parser_success}\n\n")
                
                # Статистика по сохраненным выборам
                user_selection_used = sum(1 for r in self.results if r.get('selector_used') and 'user_selection' in r.get('status', '').lower())
                if user_selection_used > 0:
                    f.write("СОХРАНЕННЫЕ ВЫБОРЫ ПОЛЬЗОВАТЕЛЯ:\n")
                    f.write(f"  Использовано: {user_selection_used}\n\n")
                
                f.write(f"{'='*70}\n")
                f.write("ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ:\n")
                f.write(f"{'='*70}\n\n")
                
                for result in self.results:
                    f.write(f"ТОВАР #{result['index'] + 1}:\n")
                    f.write(f"  Название: {result['name']}\n")
                    f.write(f"  URL: {result['url']}\n")
                    f.write(f"  Цена: {result['price'] if result['price'] else 'Не найдена'}\n")
                    f.write(f"  Характеристика: {result.get('characteristic', 'Не найдена')}\n")
                    f.write(f"  Селектор: {result.get('selector_used', 'Не указан')}\n")
                    f.write(f"  Найденный селектор: {result.get('best_found_selector', 'Не найден')}\n")
                    f.write(f"  Режим округления: {result.get('rounding_mode', 'ceil')}\n")
                    f.write(f"  Статус: {result['status']}\n")
                    f.write(f"  Время: {result['timestamp']}\n")
                    f.write("-" * 50 + "\n\n")
            
            print(f"\n{'='*60}")
            print("ОТЧЕТ СОЗДАН:")
            print(f"  Файл: {report_file}")
            print("\nСТАТИСТИКА:")
            print(f"  Успешно: {successful} из {total}")
            print(f"  Процент успеха: {(successful/total*100 if total > 0 else 0):.1f}%")
            print(f"  Режим округления: {self.rounding_mode}")
            print("  Новый парсер таблицы оргстекла: ВКЛЮЧЕН")
            print(f"  Использовано сохраненных выборов: {len(self.user_selections)}")
            print(f"{'='*60}")
            
        except Exception as e:
            logger.error(f"Ошибка генерации отчета: {e}")
    
    def set_rounding_mode_menu(self):
        """Меню установки режима округления"""
        print(f"\n{'='*60}")
        print("НАСТРОЙКА РЕЖИМА ОКРУГЛЕНИЯ ЦЕН")
        print(f"{'='*60}")
        print(f"Текущий режим: {self.rounding_mode}")
        print("\nДоступные режимы:")
        print("1. ceil - округление в большую сторону (по умолчанию)")
        print("2. floor - округление в меньшую сторону")
        print("3. nearest - округление до ближайшего целого числа")
        print("4. nearest_decimal - округление до 2 знаков после запятой")
        print("5. no_decimal - округление до целого числа в большую сторону")
        
        choice = input("\nВыберите режим (1-5, или Enter для отмены): ").strip()
        
        modes = {
            '1': 'ceil',
            '2': 'floor',
            '3': 'nearest',
            '4': 'nearest_decimal',
            '5': 'no_decimal'
        }
        
        if choice in modes:
            if self.set_rounding_mode(modes[choice]):
                print(f"\n✓ Режим округления изменен на: {modes[choice]}")
        else:
            print("Отмена")
    
    def save_selectors_to_excel(self):
        """Принудительное сохранение всех селекторов из DataFrame в Excel"""
        try:
            if not self.load_excel_data():
                print("Не удалось загрузить данные из Excel")
                return
            
            print("\nПринудительное сохранение селекторов и характеристик...")
            
            # Открываем файл для записи
            wb = load_workbook(self.excel_file)
            
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
            else:
                ws = wb[self.sheet_name]
            
            # Определяем индексы колонок
            selector_col_idx = None
            char_col_idx = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    if 'селектор' in str(cell_value).lower():
                        selector_col_idx = col
                    elif 'характеристика' in str(cell_value).lower():
                        char_col_idx = col
            
            # Если колонки "Характеристика" нет, добавляем ее
            if char_col_idx is None:
                char_col_idx = ws.max_column + 1
                ws.cell(row=1, column=char_col_idx).value = "Характеристика"
            
            if selector_col_idx:
                # Обновляем селекторы и характеристики из DataFrame
                for idx in range(len(self.df)):
                    row_idx = idx + 2
                    selector = safe_str(self.df.iloc[idx].get('Селектор', ''))
                    characteristic = safe_str(self.df.iloc[idx].get('Характеристика', ''))
                    ws.cell(row=row_idx, column=selector_col_idx).value = selector
                    ws.cell(row=row_idx, column=char_col_idx).value = characteristic
                
                # Сохраняем файл
                wb.save(self.excel_file)
                print(f"✓ Селекторы и характеристики сохранены в Excel файл: {self.excel_file}")
            else:
                print("Не удалось найти колонку 'Селектор'")
                
        except Exception as e:
            print(f"Ошибка сохранения селекторов и характеристик: {e}")
    
    def view_user_selections(self):
        """Просмотр сохраненных выборов пользователя"""
        try:
            print(f"\n{'='*60}")
            print("СОХРАНЕННЫЕ ВЫБОРЫ ПОЛЬЗОВАТЕЛЯ")
            print(f"{'='*60}")
            
            if not self.user_selections:
                print("Нет сохраненных выборов")
                return
            
            print(f"Всего сохранено выборов: {len(self.user_selections)}")
            print("\nСписок сохраненных выборов:")
            print("-" * 100)
            
            for i, (key, selection) in enumerate(self.user_selections.items(), 1):
                print(f"{i}. URL: {selection['url'][:80]}...")
                print(f"   Селектор: {selection['selector']}")
                print(f"   Текст: {selection['text'][:100]}...")
                print(f"   Цена: {selection['price']}")
                print(f"   Время: {selection['timestamp']}")
                print("-" * 100)
            
            # Предлагаем удалить выбор
            delete_choice = input("\nУдалить сохраненный выбор? (введите номер или 0 для отмены): ").strip()
            
            if delete_choice.isdigit():
                idx = int(delete_choice) - 1
                if idx >= 0:
                    keys = list(self.user_selections.keys())
                    if idx < len(keys):
                        key_to_delete = keys[idx]
                        del self.user_selections[key_to_delete]
                        self.save_user_selections()
                        print(f"✓ Выбор #{delete_choice} удален")
                    else:
                        print("Неверный номер")
            
            input("\nНажмите Enter для возврата в меню...")
            
        except Exception as e:
            print(f"Ошибка: {e}")
            traceback.print_exc()


def main():
    """Основная функция"""
    print("="*60)
    print("ПАРСЕР ЦЕН ДЛЯ EXCEL ФАЙЛА")
    print("="*60)
    print("УЛУЧШЕННАЯ ВЕРСИЯ С УНИВЕРСАЛЬНЫМ ПАРСЕРОМ TABLES")
    print("="*60)
    print("Функционал:")
    print("1. УНИВЕРСАЛЬНЫЙ парсер таблиц для Bestly.ru")
    print("2. Работает с ЛЮБЫМИ товарами: поликарбонат, оргстекло, пластик, металл и т.д.")
    print("3. Автоматическое извлечение параметров: толщина, цвет, размер, бренд, тип")
    print("4. Интеллектуальный поиск по соответствию параметров")
    print("5. Альтернативный поиск в карточках и списках товаров")
    print("6. Проверка разумности цен")
    print("="*60)
    
    excel_file = "Цены на материалы(автоматические).xlsx"
    sheet_name = "Прайс-лист"
    
    if not os.path.exists(excel_file):
        print(f"Файл {excel_file} не найден!")
        return
    
    parser = PriceParserWithSheets(excel_file, sheet_name)
    
    while True:
        print("\n" + "="*60)
        print("МЕНЮ:")
        print("1. Парсить все товары (универсальный парсер для Bestly)")
        print("2. Парсить с определенной позиции")
        print("3. Найти и выбрать селектор для товара")
        print("4. Тестировать селектор для товара")
        print("5. Проверить Selenium")
        print("6. Тестировать УНИВЕРСАЛЬНЫЙ парсер таблиц")  # НОВЫЙ ПУНКТ
        print("7. Редактировать селекторы и характеристики вручную")
        print("8. Настроить режим округления цен")
        print("9. Принудительно сохранить все селекторы и характеристики")
        print("10. Просмотреть сохраненные выборы пользователя")
        print("11. Выход")
        print("="*60)
        
        choice = input("\nВаш выбор (1-11): ").strip()
        
        if choice == '1':
            print(f"\nТекущий режим округления: {parser.rounding_mode}")
            print("ВАЖНО: Будут сохранены все ранее выбранные селекторы")
            print("Для Bestly используется Selenium с улучшенной поддержкой")
            print("Для оргстекла на Bestly.ru используется НОВЫЙ парсер таблиц")
            print("Используются СОХРАНЕННЫЕ ВЫБОРЫ ПОЛЬЗОВАТЕЛЯ")
            print("Поиск осуществляется по селектору и названию товара")
            print("Найденные характеристики сохраняются в столбец 'Характеристика'")
            confirm = input("Запустить парсинг всех товаров? (да/нет): ").strip().lower()
            if confirm in ['да', 'yes', 'y']:
                if parser.parse_all_products():
                    print("\n✓ Парсинг завершен! Селекторы и характеристики сохранены.")
                else:
                    print("\n✗ Ошибка парсинга!")
        
        elif choice == '2':
            if parser.load_excel_data():
                print(f"\nВсего товаров: {len(parser.df)}")
                print(f"Текущий режим округления: {parser.rounding_mode}")
                start = input("Начать с позиции (начиная с 1): ").strip()
                start = int(start) - 1 if start.isdigit() else 0
                
                limit = input("Сколько товаров (оставьте пустым для всех): ").strip()
                limit = int(limit) if limit.isdigit() else None
                
                print(f"\nНачинаем с позиции {start+1}...")
                parser.parse_all_products(start_from=start, limit=limit)
        
        elif choice == '3':
            # Найти и выбрать селектор для товара
            parser.find_and_select_selector_for_product()
        
        elif choice == '4':
            # Тестирование селектора для товара
            parser.test_selector_for_product()
        
        elif choice == '5':
            print("\nПроверка Selenium...")
            if parser.init_selenium_driver():
                print("✓ Selenium драйвер успешно инициализирован")
                test_url = "https://bestly.ru/catalog/org_steklo_plazcryl.html"
                print(f"Пробуем загрузить тестовую страницу: {test_url}")
                try:
                    parser.driver.get(test_url)
                    time.sleep(5)
                    print(f"✓ Страница успешно загружена: {parser.driver.title}")
                    print("Размер HTML:", len(parser.driver.page_source))
                    
                    # Пробуем найти таблицу
                    tables = parser.driver.find_elements(By.TAG_NAME, 'table')
                    print(f"Найдено таблиц на странице: {len(tables)}")
                    
                    if tables:
                        # Выводим первую таблицу
                        table_html = tables[0].get_attribute('outerHTML')
                        soup = BeautifulSoup(table_html, 'html.parser')
                        rows = soup.find_all('tr')
                        print(f"В первой таблице строк: {len(rows)}")
                        
                        # Выводим первые 3 строки
                        for i, row in enumerate(rows[:4]):
                            cells = row.find_all(['td', 'th'])
                            cell_texts = [cell.get_text(strip=True) for cell in cells]
                            print(f"Строка {i}: {cell_texts}")
                    
                    parser.close_selenium_driver()
                    print("✓ Драйвер успешно закрыт")
                except Exception as e:
                    print(f"✗ Ошибка загрузки страницы: {e}")
                    parser.close_selenium_driver()
            else:
                print("✗ Не удалось инициализировать Selenium драйвер")
        
        elif choice == '6':
            # Тестирование универсального парсера табли
            parser.test_universal_parser()

        elif choice == '7':
            print("\nРедактирование селекторов и характеристик вручную")
            if parser.load_excel_data():
                print(f"\nВсего товаров: {len(parser.df)}")
                print("Формат: номер - название - текущий селектор - характеристика")
                print("-" * 80)
                
                for i in range(min(20, len(parser.df))):
                    row = parser.df.iloc[i]
                    # Используем safe_str для безопасного преобразования
                    name = safe_str(row.get('Название'), f'Товар {i+1}')
                    selector = safe_str(row.get('Селектор', ''))
                    characteristic = safe_str(row.get('Характеристика', ''))
                    print(f"{i+1:3}. {name[:30]}... | Селектор: {selector} | Характеристика: {characteristic[:30]}...")
                
                print("-" * 80)
                
                edit_choice = input("\nВведите номер товара для редактирования (0 для отмены): ").strip()
                if edit_choice.isdigit():
                    idx = int(edit_choice) - 1
                    if 0 <= idx < len(parser.df):
                        row = parser.df.iloc[idx]
                        current_selector = safe_str(row.get('Селектор', ''))
                        current_char = safe_str(row.get('Характеристика', ''))
                        name = safe_str(row.get('Название'), f'Товар {idx+1}')
                        
                        print(f"\nРедактирование товара #{idx+1}: {name}")
                        print(f"Текущий селектор: {current_selector}")
                        print(f"Текущая характеристика: {current_char}")
                        
                        new_selector = input("\nВведите новый селектор (оставьте пустым для удаления): ").strip()
                        new_char = input("Введите новую характеристику (оставьте пустым для удаления): ").strip()
                        
                        if new_selector == "" and new_char == "":
                            print("Ничего не изменено")
                        else:
                            # Обновляем в DataFrame
                            if new_selector != "":
                                parser.df.at[idx, 'Селектор'] = new_selector
                            if new_char != "":
                                parser.df.at[idx, 'Характеристика'] = new_char
                            
                            # Сохраняем в Excel
                            try:
                                wb = load_workbook(parser.excel_file)
                                if parser.sheet_name in wb.sheetnames:
                                    ws = wb[parser.sheet_name]
                                    
                                    # Находим колонки
                                    sel_col = None
                                    char_col = None
                                    for col in range(1, ws.max_column + 1):
                                        header = ws.cell(row=1, column=col).value
                                        if header:
                                            if 'селектор' in str(header).lower():
                                                sel_col = col
                                            elif 'характеристика' in str(header).lower():
                                                char_col = col
                                    
                                    # Если колонки "Характеристика" нет, добавляем ее
                                    if char_col is None:
                                        char_col = ws.max_column + 1
                                        ws.cell(row=1, column=char_col).value = "Характеристика"
                                    
                                    # Обновляем значения
                                    row_idx = idx + 2
                                    if sel_col and new_selector != "":
                                        ws.cell(row=row_idx, column=sel_col).value = new_selector
                                    if char_col and new_char != "":
                                        ws.cell(row=row_idx, column=char_col).value = new_char
                                    
                                    wb.save(parser.excel_file)
                                    print("✓ Изменения сохранены в Excel")
                                else:
                                    print("Лист не найден")
                            except Exception as e:
                                print(f"Ошибка сохранения: {e}")
                    else:
                        print("Неверный номер товара")
        
        elif choice == '8':
            # Настроить режим округления цен
            parser.set_rounding_mode_menu()
        
        elif choice == '9':
            # Принудительное сохранение всех селекторов и характеристик
            parser.save_selectors_to_excel()
        
        elif choice == '10':
            # Просмотр сохраненных выборов пользователя
            parser.view_user_selections()
        
        elif choice == '11':
            print("\nВыход.")
            break


if __name__ == "__main__":
    main()
