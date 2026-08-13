"""
Юнит-тесты для price_parser_sheets.py.

Покрывают чистую логику (извлечение цены/толщины/цвета из текста,
сопоставление толщины, парсинг табличных цен из HTML, стабильность
ключа сохраненного выбора пользователя) без обращения к реальным
сайтам и без запуска настоящего браузера.

Selenium и webdriver_manager не обязательны для запуска этих тестов:
если они не установлены, вместо них подставляются заглушки-модули,
поскольку price_parser_sheets.py импортирует их на уровне модуля, но
код, покрытый тестами ниже, их не использует.

Запуск:
    python -m unittest discover -s tests
"""
import importlib.util
import os
import sys
import tempfile
import types
import unittest
from unittest.mock import patch, MagicMock

from openpyxl import Workbook, load_workbook


def _install_selenium_stubs():
    """Подставляет минимальные заглушки selenium/webdriver_manager,
    если они не установлены, чтобы модуль можно было импортировать."""
    try:
        import selenium  # noqa: F401
        return
    except ImportError:
        pass

    for mod_name in [
        'selenium', 'selenium.webdriver', 'selenium.webdriver.common.by',
        'selenium.webdriver.support.ui', 'selenium.webdriver.support',
        'selenium.webdriver.chrome.service', 'selenium.webdriver.chrome.options',
        'selenium.common.exceptions', 'webdriver_manager', 'webdriver_manager.chrome',
    ]:
        sys.modules[mod_name] = types.ModuleType(mod_name)

    sys.modules['selenium'].webdriver = types.SimpleNamespace(Chrome=object)
    sys.modules['selenium.webdriver.common.by'].By = types.SimpleNamespace(
        CSS_SELECTOR='css selector', XPATH='xpath', TAG_NAME='tag name'
    )
    sys.modules['selenium.webdriver.support.ui'].WebDriverWait = object
    sys.modules['selenium.webdriver.support'].expected_conditions = types.SimpleNamespace()
    sys.modules['selenium.webdriver.chrome.service'].Service = object
    sys.modules['selenium.webdriver.chrome.options'].Options = object
    sys.modules['selenium.common.exceptions'].TimeoutException = Exception
    sys.modules['webdriver_manager.chrome'].ChromeDriverManager = object


_install_selenium_stubs()

_MODULE_PATH = os.path.join(os.path.dirname(__file__), '..', 'price_parser_sheets.py')
_spec = importlib.util.spec_from_file_location('price_parser_sheets', _MODULE_PATH)
pps = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(pps)


def make_parser():
    p = pps.PriceParserWithSheets.__new__(pps.PriceParserWithSheets)
    p.rounding_mode = 'ceil'
    return p


def make_workbook(tmpdir, name, url, selector=""):
    """Creates a one-row test Прайс-лист workbook and returns its path."""
    xlsx_path = os.path.join(tmpdir, "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    ws.append(["Название", "URL", "Цена", "Селектор", "Характеристика", "Дата обновления"])
    ws.append([name, url, "", selector, "", ""])
    wb.save(xlsx_path)
    return xlsx_path


def make_fake_response(url, text, status_code=200):
    resp = MagicMock()
    resp.status_code = status_code
    resp.url = url
    resp.encoding = 'utf-8'
    resp.text = text
    return resp


def make_parser_for_product(tmpdir, name, url, selector=""):
    """Convenience: build a workbook + loaded parser in one call."""
    xlsx_path = make_workbook(tmpdir, name, url, selector=selector)
    p = pps.PriceParserWithSheets(xlsx_path, sheet_name="Прайс-лист")
    p.load_excel_data()
    return p


def no_selenium():
    """
    bestly.ru always tries Selenium first in parse_single_product(). Without
    this patch, tests using a bestly.ru URL would make a real
    webdriver_manager network call (to look up a chromedriver version)
    before falling back to the mocked requests.get — slow, and a violation
    of this suite's "no network needed" guarantee. Forces that fallback
    immediately instead.
    """
    return patch.object(pps.PriceParserWithSheets, 'get_with_selenium', return_value=None)


class SafeStrTests(unittest.TestCase):
    def test_none_returns_default(self):
        self.assertEqual(pps.safe_str(None), '')

    def test_nan_returns_default(self):
        self.assertEqual(pps.safe_str(float('nan')), '')

    def test_regular_value(self):
        self.assertEqual(pps.safe_str('hello'), 'hello')
        self.assertEqual(pps.safe_str(123), '123')


class ExtractPriceFromTextTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_price_with_currency_and_thousands_separator(self):
        self.assertEqual(self.p.extract_price_from_text("6 388.80₽"), 6388.80)

    def test_price_with_comma_decimal(self):
        self.assertEqual(self.p.extract_price_from_text("от 1 234,56 руб."), 1234.56)

    def test_plain_integer(self):
        self.assertEqual(self.p.extract_price_from_text("7686"), 7686.0)

    def test_price_with_abbreviation(self):
        self.assertEqual(self.p.extract_price_from_text("Цена: 999 р."), 999.0)

    def test_none_and_empty(self):
        self.assertIsNone(self.p.extract_price_from_text(None))
        self.assertIsNone(self.p.extract_price_from_text(""))


class ExtractThicknessTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_thickness_with_mm_suffix(self):
        self.assertEqual(
            self.p.extract_thickness_from_product_name("Орг.стекло PLAZCRYL прозрачный 4мм"),
            "4.0",
        )

    def test_thickness_before_color(self):
        self.assertEqual(
            self.p.extract_thickness_from_product_name("Орг.стекло 8мм прозрачный"),
            "8.0",
        )


class ExtractColorMaterialBrandTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_color_direct_match(self):
        self.assertEqual(
            self.p.extract_color_from_product_name("Орг.стекло PLAZCRYL прозрачный 4мм"),
            "прозрачный",
        )

    def test_color_variation_match(self):
        self.assertEqual(
            self.p.extract_color_from_product_name("Орг.стекло матовое 4мм"),
            "матовый",
        )

    def test_material_type(self):
        self.assertEqual(
            self.p.extract_material_type_from_product_name("Поликарбонат POLYGAL сотовый 4мм"),
            "поликарбонат",
        )

    def test_brand_uppercase(self):
        self.assertEqual(
            self.p.extract_brand_from_product_name("Орг.стекло PLAZCRYL прозрачный 4мм"),
            "PLAZCRYL",
        )

    def test_product_type(self):
        self.assertEqual(
            self.p.extract_product_type_from_product_name("Поликарбонат сотовый 4мм"),
            "сотовый",
        )


class RoundPriceTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_ceil_mode(self):
        self.assertEqual(self.p.round_price(10.001, mode='ceil'), 10.01)

    def test_floor_mode(self):
        self.assertEqual(self.p.round_price(10.999, mode='floor'), 10.99)

    def test_no_decimal_mode(self):
        self.assertEqual(self.p.round_price(10.01, mode='no_decimal'), 11)


class IsReasonablePriceTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_within_configured_range(self):
        self.assertTrue(self.p.is_reasonable_price(7686, "Поликарбонат POLYGAL сотовый 4мм"))

    def test_outside_configured_range(self):
        self.assertFalse(self.p.is_reasonable_price(500, "Поликарбонат POLYGAL сотовый 4мм"))

    def test_unknown_product_uses_wide_range(self):
        self.assertTrue(self.p.is_reasonable_price(500, "какой-то неизвестный товар"))

    def test_non_positive_price_rejected(self):
        self.assertFalse(self.p.is_reasonable_price(-1, "что угодно"))
        self.assertFalse(self.p.is_reasonable_price(0, "что угодно"))


class ThicknessMatchingTests(unittest.TestCase):
    """Regression tests for the thickness-matching helpers (dedup + bugfix stages)."""

    def setUp(self):
        self.p = make_parser()

    def test_matches_exact_mm_mention(self):
        self.assertTrue(self.p.check_thickness_in_text("Толщина: 4мм, цена 500р", "4"))

    def test_does_not_match_different_thickness(self):
        self.assertFalse(self.p.check_thickness_in_text("6мм лист", "4"))

    def test_no_false_positive_on_substring_of_longer_number(self):
        # Regression test: "4" must not match inside "14мм"/"24мм"/"144мм".
        self.assertFalse(self.p.check_thickness_in_text("14мм что-то не то", "4"))
        self.assertFalse(self.p.check_thickness_in_text("24мм странная толщина", "4"))
        self.assertFalse(self.p.check_thickness_in_text("144мм странное совпадение", "4"))

    def test_exact_mode_requires_whole_cell(self):
        self.assertTrue(self.p._text_matches_thickness("4", "4.0", mode='exact'))
        self.assertFalse(self.p._text_matches_thickness("14", "4.0", mode='exact'))

    def test_find_best_match_by_thickness_picks_correct_candidate(self):
        elements = [
            {'text': '4мм прозрачный 1234 руб.'},
            {'text': '6мм прозрачный 2345 руб.'},
        ]
        best = self.p.find_best_match_by_thickness(elements, "4.0")
        self.assertIsNotNone(best)
        self.assertEqual(best['price'], 1234.0)


class BestlyTableParsingTests(unittest.TestCase):
    """Parse synthetic HTML tables shaped like the real bestly.ru pages."""

    def setUp(self):
        self.p = make_parser()

    def test_universal_table_parsing_bestly(self):
        html = """
        <html><body>
        <table>
        <tr><th>Толщина</th><th>Размер</th><th>Цена</th></tr>
        <tr><td>4 мм</td><td>2050x3050</td><td>7 686 ₽</td></tr>
        <tr><td>6 мм</td><td>2050x3050</td><td>10 500 ₽</td></tr>
        </table>
        </body></html>
        """
        price = self.p.universal_table_parsing_bestly(html, "Поликарбонат POLYGAL сотовый 4мм")
        self.assertEqual(price, 7686.0)

    def test_parse_orgsteklo_table_improved(self):
        html = """
        <html><body>
        <table>
        <tr><td>4мм</td><td>2050x3050</td><td>12 345.00</td></tr>
        <tr><td>6мм</td><td>2050x3050</td><td>15 000.00</td></tr>
        </table>
        </body></html>
        """
        price = self.p.parse_orgsteklo_table_improved(html, "Орг.стекло PLAZCRYL прозрачный 4мм")
        self.assertEqual(price, 12345.0)

    def test_parse_bestly_orgsteklo_table(self):
        html = """
        <html><body>
        <table>
        <tr><td></td><td>цвет</td><td>толщина</td><td>размер</td><td>цена</td></tr>
        <tr><td><input></td><td>прозрачный</td><td>4</td><td>2050x3050</td><td>9 999 ₽</td></tr>
        <tr><td><input></td><td>прозрачный</td><td>6</td><td>2050x3050</td><td>13 000 ₽</td></tr>
        </table>
        </body></html>
        """
        price = self.p.parse_bestly_orgsteklo_table(html, "Орг.стекло PLAZCRYL прозрачный 4мм")
        self.assertEqual(price, 9999.0)


class UserSelectionPersistenceTests(unittest.TestCase):
    """Regression test for the hash()-instability bugfix (stage 1)."""

    def setUp(self):
        self._old_cwd = os.getcwd()
        self._tmpdir = tempfile.mkdtemp()
        os.chdir(self._tmpdir)

    def tearDown(self):
        os.chdir(self._old_cwd)

    def test_same_selection_produces_stable_key_across_instances(self):
        p1 = make_parser()
        p1.user_selections = {}
        p1.save_user_selection("https://bestly.ru/x.html", ".price", "1234 руб.", 1234.0)
        key1 = list(p1.user_selections.keys())[0]

        p2 = make_parser()
        p2.user_selections = {}
        p2.load_user_selections()
        p2.save_user_selection("https://bestly.ru/x.html", ".price", "1234 руб.", 1234.0)
        key2 = list(p2.user_selections.keys())[0]

        self.assertEqual(key1, key2)
        self.assertEqual(len(p2.user_selections), 1)


class ExtractDomainTests(unittest.TestCase):
    def setUp(self):
        self.p = make_parser()

    def test_strips_www_and_scheme(self):
        self.assertEqual(self.p.extract_domain("https://www.bestly.ru/catalog/x.html"), "bestly.ru")

    def test_empty_for_invalid_input(self):
        self.assertEqual(self.p.extract_domain(None), '')


class StripTrackingParamsTests(unittest.TestCase):
    def test_removes_ysclid(self):
        self.assertEqual(
            pps.strip_tracking_params("https://expo-torg.ru/product/x/?ysclid=mgz4hztunr807680454"),
            "https://expo-torg.ru/product/x/",
        )

    def test_removes_utm_params_keeps_real_ones(self):
        self.assertEqual(
            pps.strip_tracking_params("https://site.ru/p/?utm_source=yandex&utm_medium=cpc&real=1"),
            "https://site.ru/p/?real=1",
        )

    def test_keeps_legitimate_query_param(self):
        self.assertEqual(
            pps.strip_tracking_params("https://site.ru/p/?id=123&ysclid=abc"),
            "https://site.ru/p/?id=123",
        )

    def test_no_query_string_unchanged(self):
        self.assertEqual(pps.strip_tracking_params("https://site.ru/p/"), "https://site.ru/p/")

    def test_empty_and_none(self):
        self.assertEqual(pps.strip_tracking_params(""), "")
        self.assertIsNone(pps.strip_tracking_params(None))


class GetWithRequestsUrlHealingTests(unittest.TestCase):
    """Self-healing of a product URL when the site redirects to a new address."""

    def setUp(self):
        self.p = make_parser()

    def test_captures_redirect_target(self):
        fake_response = MagicMock()
        fake_response.status_code = 200
        fake_response.url = "https://expo-torg.ru/catalog/ldsp/new-slug/"
        fake_response.encoding = 'utf-8'
        fake_response.text = "<html>ok</html>"

        with patch.object(pps.requests, 'get', return_value=fake_response):
            html = self.p.get_with_requests("https://expo-torg.ru/catalog/ldsp/old-slug/")

        self.assertEqual(html, "<html>ok</html>")
        self.assertEqual(self.p.last_status_code, 200)
        self.assertEqual(self.p.last_fetched_url, "https://expo-torg.ru/catalog/ldsp/new-slug/")

    def test_no_redirect_leaves_url_unchanged(self):
        fake_response = MagicMock()
        fake_response.status_code = 200
        fake_response.url = "https://expo-torg.ru/catalog/ldsp/same/"
        fake_response.encoding = 'utf-8'
        fake_response.text = "<html>ok</html>"

        with patch.object(pps.requests, 'get', return_value=fake_response):
            self.p.get_with_requests("https://expo-torg.ru/catalog/ldsp/same/")

        self.assertEqual(self.p.last_fetched_url, "https://expo-torg.ru/catalog/ldsp/same/")

    def test_404_is_recorded_and_returns_none(self):
        fake_response = MagicMock()
        fake_response.status_code = 404

        with patch.object(pps.requests, 'get', return_value=fake_response):
            html = self.p.get_with_requests("https://expo-torg.ru/catalog/ldsp/gone/")

        self.assertIsNone(html)
        self.assertEqual(self.p.last_status_code, 404)
        self.assertIsNone(self.p.last_fetched_url)


class ParseSingleProductRegressionTests(unittest.TestCase):
    """
    Regression tests for a real bug found while testing the URL self-heal
    feature: the "standard search" block in parse_single_product() was
    nested entirely inside `if selector and selector.strip():`, so:
      1. a product with NO selector set (the common case before a selector
         is manually chosen) crashed with UnboundLocalError on `result`
         instead of searching for the price at all;
      2. a product whose specified selector failed, but whose price WAS
         found via the bestly.ru-specific fallback selectors, had that
         price silently wiped back to None/"Цена не найдена" by a
         duplicated, mis-nested copy of the same search block.
    Both cases are exercised here with mocked HTTP responses (no network).
    """

    def test_no_selector_does_not_crash_and_finds_price(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Композитная панель белая 4мм",
                "https://bestly.ru/catalog/composite/x.html", selector="",
            )
            fake_response = make_fake_response(
                "https://bestly.ru/catalog/composite/x.html",
                "<html><body><div class='item_price'>5000 руб.</div></body></html>",
            )
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 5000.0)
        self.assertNotIn('Ошибка', result['status'])

    def test_wrong_selector_falls_back_without_wiping_found_price(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Композитная панель белая 4мм",
                "https://bestly.ru/catalog/composite/x.html",
                selector=".wrong-selector-does-not-exist",
            )
            fake_response = make_fake_response(
                "https://bestly.ru/catalog/composite/x.html",
                "<html><body><div class='item_price'>5000 руб.</div></body></html>",
            )
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 5000.0)

    def test_correct_selector_still_works(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Композитная панель белая 4мм",
                "https://bestly.ru/catalog/composite/x.html", selector=".item_price",
            )
            fake_response = make_fake_response(
                "https://bestly.ru/catalog/composite/x.html",
                "<html><body><div class='item_price'>5000 руб.</div></body></html>",
            )
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 5000.0)
        self.assertIn('указанный селектор', result['status'])

    def test_nothing_found_is_a_clean_status_not_a_crash(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "ЛДСП белый 16мм", "https://expo-torg.ru/catalog/ldsp/y.html", selector="",
            )
            fake_response = make_fake_response(
                "https://expo-torg.ru/catalog/ldsp/y.html",
                "<html><body><p>ничего интересного</p></body></html>",
            )
            with patch.object(pps.requests, 'get', return_value=fake_response):
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertIsNone(result['price'])
        self.assertNotIn('Ошибка', result['status'])

    def test_redirected_url_is_resolved_and_written_back_to_excel(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "ЛДСП белый 16мм",
                "https://expo-torg.ru/catalog/ldsp/old-slug/?ysclid=abc123", selector="",
            )
            fake_response = make_fake_response(
                "https://expo-torg.ru/catalog/ldsp/new-slug-16mm/",
                "<html><body><div class='price'>1234 руб.</div></body></html>",
            )
            with patch.object(pps.requests, 'get', return_value=fake_response):
                result = p.parse_single_product(0, p.df.iloc[0])

            self.assertEqual(result['url'], "https://expo-torg.ru/catalog/ldsp/new-slug-16mm/")

            p.results = [result]
            self.assertTrue(p.save_results_to_excel())

            wb2 = load_workbook(p.excel_file)
            saved_url = wb2["Прайс-лист"].cell(row=2, column=2).value
            self.assertEqual(saved_url, "https://expo-torg.ru/catalog/ldsp/new-slug-16mm/")

    def test_empty_result_url_does_not_blank_existing_cell(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "test.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Прайс-лист"
            ws.append(["Название", "URL", "Цена", "Селектор", "Характеристика", "Дата обновления"])
            ws.append(["Товар без URL", "", "", "", "", ""])
            wb.save(xlsx_path)

            p = pps.PriceParserWithSheets(xlsx_path, sheet_name="Прайс-лист")
            p.results = [{
                'index': 0, 'name': 'Товар без URL', 'url': '', 'price': None,
                'characteristic': '', 'selector_used': '', 'best_found_selector': '',
                'status': 'URL не указан', 'rounding_mode': 'ceil',
                'timestamp': '2026-01-01 00:00:00',
            }]
            self.assertTrue(p.save_results_to_excel())

            wb2 = load_workbook(xlsx_path)
            cell = wb2["Прайс-лист"].cell(row=2, column=2).value
            self.assertIn(cell, (None, ''))


class WrongPriceRegressionTests(unittest.TestCase):
    """
    Regression tests for real "wrong price" cases reported from a live
    bestly.ru parsing run:
      1. The generic '[data-price]' fallback selector sometimes matches a
         quantity-stepper widget (data-price="1") instead of the actual
         price element, so "1" was accepted as the price outright.
      2. A "soft 404" — the site returns HTTP 200 but the page content is
         an actual "not found" page — had its title/h1 text ("Страница не
         найдена (404 Not Found)") auto-detected as a price of 404.
    Both are fixed by (a) sanity-checking every extracted price with
    is_reasonable_price() before accepting it in
    find_price_with_selector_and_name()/find_price_and_name_on_page(), and
    (b) looks_like_error_page() short-circuiting parse_single_product()
    before any price search runs.
    """

    def test_data_price_quantity_widget_is_rejected_in_favor_of_real_price(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Иглопробивной ковролин EXPORADU",
                "https://bestly.ru/catalog/exporadu.html", selector="",
            )
            html = (
                "<html><body>"
                "<div class='quantity-stepper' data-price='1'>шт.</div>"
                "<div class='item_price'>2450 руб.</div>"
                "</body></html>"
            )
            fake_response = make_fake_response("https://bestly.ru/catalog/exporadu.html", html)
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 2450.0)

    def test_data_price_one_alone_is_reported_as_not_found_not_as_price_one(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Товар без реальной цены на странице",
                "https://bestly.ru/catalog/nothing_real.html", selector="",
            )
            html = "<html><body><div class='quantity-stepper' data-price='1'>шт.</div></body></html>"
            fake_response = make_fake_response("https://bestly.ru/catalog/nothing_real.html", html)
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertIsNone(result['price'])

    def test_plausible_data_price_still_works(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Товар с ценой в data-price",
                "https://bestly.ru/catalog/real_dataprice.html", selector="",
            )
            html = "<html><body><div data-price='3500'>3500 руб.</div></body></html>"
            fake_response = make_fake_response("https://bestly.ru/catalog/real_dataprice.html", html)
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 3500.0)

    def test_soft_404_page_is_not_mistaken_for_a_price(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Баннерная ткань литая Blackback GLP",
                "https://bestly.ru/catalog/bannernaya_tkan_litaya_blackback_glp.html", selector="",
            )
            html = (
                "<html><head><title>Страница не найдена (404 Not Found)</title></head>"
                "<body><h1>Страница не найдена (404 Not Found)</h1></body></html>"
            )
            fake_response = make_fake_response(
                "https://bestly.ru/catalog/bannernaya_tkan_litaya_blackback_glp.html", html,
            )
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertIsNone(result['price'])
        self.assertIn('заглушка', result['status'])

    def test_sku_containing_404_does_not_false_positive(self):
        """A product whose SKU happens to contain '404' must still parse normally."""
        with tempfile.TemporaryDirectory() as tmpdir:
            p = make_parser_for_product(
                tmpdir, "Артикул 404XYZ товар", "https://bestly.ru/catalog/sku404.html", selector="",
            )
            html = (
                "<html><head><title>Товар 404XYZ - каталог</title></head>"
                "<body><h1>Товар 404XYZ</h1><div class='item_price'>999 руб.</div></body></html>"
            )
            fake_response = make_fake_response("https://bestly.ru/catalog/sku404.html", html)
            with patch.object(pps.requests, 'get', return_value=fake_response), no_selenium():
                result = p.parse_single_product(0, p.df.iloc[0])

        self.assertEqual(result['price'], 999.0)

    def test_looks_like_error_page_direct(self):
        p = make_parser()
        self.assertTrue(p.looks_like_error_page(
            "<html><head><title>Страница не найдена</title></head><body></body></html>"
        ))
        self.assertFalse(p.looks_like_error_page(
            "<html><head><title>Товар XYZ</title></head><body><div class='price'>100</div></body></html>"
        ))
        self.assertFalse(p.looks_like_error_page(""))
        self.assertFalse(p.looks_like_error_page(None))


if __name__ == '__main__':
    unittest.main()
